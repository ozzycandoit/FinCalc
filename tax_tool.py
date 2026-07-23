#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 TURKEY FOREIGN STOCK / ETF CAPITAL-GAINS TOOL - expandable edition
================================================================================

Commands:
  python tax_tool.py init [--assets N] [--rows N] [--overwrite]
      Creates Turkey_Tax_Tracker.xlsx with a SUMMARY sheet and N ASSET sheets.
      Default: 10 assets and 500 transaction rows per asset.

  python tax_tool.py add-asset [N] [--rows N]
      Adds N more ASSET sheets to an existing workbook and rebuilds SUMMARY.
      Default: add 1 asset.

  python tax_tool.py run [EVDS_KEY]
      Processes every sheet whose name starts with ASSET_. The number of asset
      sheets is not fixed. The SUMMARY sheet is rebuilt dynamically.

  python tax_tool.py guide
      Writes the dividend guide text file next to the workbook.

Main capital-gains rules implemented:
  - TCMB forex buying rate is used for each transaction date. If no bulletin
    exists on that date, the tool walks back to the latest published bulletin.
  - FIFO lot matching is used for sales.
  - YI-UFE indexation uses the month before acquisition and the month before
    disposal. Indexation is applied only if the index increase is at least 10%.
  - Same-year securities gains and losses are netted across all ASSET sheets;
    the annual loss floor is applied only after the total securities result.
  - If an index month needed for a realised sale is missing, the workbook is
    marked PROVISIONAL. Non-indexed cost is used only as an estimate.

Dividends are not calculated in this file. See Dividend_Calculation_Guide.md.
"""

import datetime as dt
import hashlib
import json
import re
import sys
import xml.etree.ElementTree as ET
from itertools import groupby
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing openpyxl. Run: pip install openpyxl requests")

try:
    import requests
except ImportError:
    sys.exit("Missing requests. Run: pip install openpyxl requests")


# -----------------------------------------------------------------------------
# CONFIG
# -----------------------------------------------------------------------------
FILENAME = "Turkey_Tax_Tracker.xlsx"
TAX_YEAR = 2026
DEFAULT_ASSETS = 10
DEFAULT_MAX_TX = 500
TX_FIRST = 11

# Option sheets (one option series per sheet)
DEFAULT_OPTIONS = 5
OPT_DETAIL_FIRST = 4       # detail rows 4..10
OPT_TX_FIRST = 14          # first transaction row
OPT_DEFAULT_ROWS = 100

# How same-day sells pick which shares they dispose of:
#   False -> STRICT FIFO (default): the oldest shares you hold are deemed sold
#            first, even on a same-day round-trip. This is the literal reading
#            of GIB FIFO ("ilk giren ilk cikar").
#   True  -> SAME-DAY RULE (UK-style): a same-day sell is matched against that
#            day's buys first, then the older pool. Keeps a flat intraday
#            round-trip from disposing of your long-term lot.
# Either way, the result does NOT depend on whether you type the sell or the
# buy first on a given day, and a genuine oversell is still flagged.
SAME_DAY_MATCHING = False

# 2026 progressive income-tax brackets for non-employment income.
# Tuple format: (upper_limit, rate, cumulative_tax_at_lower_edge)
BRACKETS_BY_YEAR = {
    2026: [
        (190000, 0.15, 0),
        (400000, 0.20, 28500),
        (1000000, 0.27, 70500),
        (5300000, 0.35, 232500),
        (float("inf"), 0.40, 1737500),
    ]
}

# EVDS YI-UFE series candidates. TP.TUFE1YI.T1 is kept intentionally.
EVDS_CACHE_FILE = Path(__file__).parent / ".evds_cache.json"
EVDS_YIUFE_CANDIDATES = ["TP.TUFE1YI.T1", "TP.FG.J0"]
EVDS_ENDPOINTS = [
    "https://evds2.tcmb.gov.tr/service/evds/series={series}&startDate={start}&endDate={end}&type=json",
    "https://evds3.tcmb.gov.tr/igmevdsms-dis/series={series}&startDate={start}&endDate={end}&type=json",
]

# Manual YI-UFE fallback if EVDS is not used or unavailable.
# Fill with official monthly YI-UFE values, e.g. "2026-05": 5000.00.
UFE_TABLE = {
}

TCMB_XML = "https://www.tcmb.gov.tr/kurlar/{ym}/{dmy}.xml"

# Styling constants
NAVY = "1F3864"
WHITE = "FFFFFF"
SEC = "D6E4F0"
SECFG = "1F3864"
WARN = "FFFF00"
POS = "E2EFDA"
NEG = "FCE4D6"
INP = "0000FF"
CALC = "000000"
LINK = "006400"
BUYBG = "E2EFDA"
SELLBG = "FCE4D6"
TLF = '#,##0.00;(#,##0.00);-'
FXF = '#,##0.0000'
CUR = '#,##0.0000'
PCT = '0.00%'
DATEF = 'DD/MM/YYYY'
QTYF = '#,##0.######'


def F(sz=10, b=False, c=CALC):
    return Font(name="Arial", size=sz, bold=b, color=c)


def fill(c):
    return PatternFill("solid", fgColor=c)


def AC():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def AL():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def AR():
    return Alignment(horizontal="right", vertical="center")


_THIN = Side(style="thin", color="BFBFBF")


def B():
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def sect(ws, row, c1, c2, text, bg=NAVY, fg=WHITE):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row, c1)
    cell.value = text
    cell.font = F(10, True, fg)
    cell.fill = fill(bg)
    cell.alignment = AL()
    cell.border = B()


def safe_unmerge(ws, coord):
    try:
        ws.unmerge_cells(coord)
    except ValueError:
        pass


# -----------------------------------------------------------------------------
# WORKBOOK BUILDING
# -----------------------------------------------------------------------------
def build_asset_sheet(wb, idx, tx_rows=DEFAULT_MAX_TX, lang="TR"):
    ws = wb.create_sheet(f"ASSET_{idx:02d}")
    ws.sheet_view.showGridLines = False
    widths = {
        "A": 3, "B": 5, "C": 13, "D": 11, "E": 12, "F": 13,
        "G": 13, "H": 13, "I": 15, "J": 15, "K": 15, "L": 15,
        "M": 17, "N": 17,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B1:N1")
    t = ws["B1"]
    if lang == "TR":
        t.value = f"VARLIK {idx:02d} - İŞLEMLER VE VERGİ (GVK Mük. 80/81)"
        details_header = "VARLIK BİLGİLERİ (sarı hücreleri doldurun)"
        details = [
            ("Varlık adı / ticker", ""),
            ("Para birimi (USD veya EUR)", "USD"),
            ("Borsa", "NYSE / NASDAQ / Xetra / ..."),
            ("Vergi yılı", TAX_YEAR),
        ]
        tx_header = "İŞLEMLER — tüm ALIM ve SATIM işlemlerini girin"
    else:
        t.value = f"ASSET {idx:02d} - TRANSACTIONS AND TAX (GVK Muk. 80/81)"
        details_header = "ASSET DETAILS (fill the yellow cells)"
        details = [
            ("Asset name / ticker", ""),
            ("Currency (USD or EUR)", "USD"),
            ("Exchange", "NYSE / NASDAQ / Xetra / ..."),
            ("Tax year", TAX_YEAR),
        ]
        tx_header = "TRANSACTIONS — enter every BUY and SELL in date order"
    t.font = F(12, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = AC()

    sect(ws, 3, 2, 6, details_header, SEC, SECFG)
    for i, (lbl, val) in enumerate(details, 4):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        lc = ws.cell(i, 2)
        lc.value = lbl
        lc.font = F(10, True)
        lc.alignment = AL()
        lc.border = B()
        ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)
        vc = ws.cell(i, 4)
        vc.value = val
        vc.font = F(10, False, INP)
        vc.fill = fill(WARN)
        vc.alignment = AL()
        vc.border = B()

    sect(ws, 9, 2, 14, tx_header, NAVY)
    write_asset_headers(ws, lang=lang)

    for r in range(TX_FIRST, TX_FIRST + tx_rows):
        ws.cell(r, 2).value = r - TX_FIRST + 1
        ws.cell(r, 2).font = F(9)
        ws.cell(r, 2).alignment = AC()
        ws.cell(r, 2).border = B()
        for c in range(3, 15):
            cell = ws.cell(r, c)
            cell.border = B()
            cell.alignment = AR()
            if c in (3, 4, 5, 6, 7):
                cell.fill = fill("EBF3FB")
                cell.font = F(9, False, INP)
                if c == 3:
                    cell.number_format = DATEF
                elif c == 4:
                    cell.alignment = AC()
                elif c == 5:
                    cell.number_format = QTYF
                else:
                    cell.number_format = CUR
            elif c in (8, 9):
                cell.fill = fill("FFF7E6")
                cell.font = F(9, False, CALC)
                cell.number_format = FXF if c == 8 else '#,##0.00'
            else:
                cell.fill = fill("F2F2F2")
                cell.font = F(9, False, LINK)
                cell.number_format = TLF

    rb = TX_FIRST + tx_rows + 1
    write_asset_results_block(ws, rb, lang=lang)
    ws.freeze_panes = "B11"
    return ws


def write_asset_headers(ws, lang="TR"):
    if lang == "TR":
        heads = [
            "#",
            "Tarih",
            "İşlem Türü\n(ALIM/SATIM)",
            "Adet",
            "Fiyat\n(hisse başına,\nvarlık dövizi)",
            "Komisyon\n(varlık dövizi,\nboş = 0)",
            "TCMB kuru\n(otomatik)",
            "Yİ-ÜFE\nönceki ay\n(otomatik)",
            "TL maliyet /\nhasılat\n(otomatik)",
            "Eşlenen\nmaliyet TL\n(otomatik)",
            "Endekslenmiş\nmaliyet TL\n(otomatik)",
            "Brüt sonuç\nTL\n(otomatik)",
            "Vergiye tabi\nsonuç TL\n(otomatik)",
        ]
    else:
        heads = [
            "#",
            "Date",
            "Type\n(BUY/SELL)",
            "Quantity",
            "Price\n(per share,\nasset ccy)",
            "Commission\n(asset ccy,\nblank = 0)",
            "TCMB rate\n(auto)",
            "YI-UFE index\nprev month\n(auto)",
            "TL cost /\nproceeds\n(auto)",
            "Matched\ncost TL\n(auto)",
            "Indexed\ncost TL\n(auto)",
            "Gross result\nTL\n(auto)",
            "Taxable result\nTL\n(auto)",
        ]
    for ci, h in enumerate(heads, 2):
        c = ws.cell(10, ci)
        c.value = h
        c.font = F(9, True, WHITE)
        c.fill = fill(NAVY)
        c.alignment = AC()
        c.border = B()


def write_asset_results_block(ws, rb, lang="TR"):
    if lang == "TR":
        sect(ws, rb, 2, 14, "BU VARLIK İÇİN SONUÇLAR (hesaplama sonrası otomatik doldurulur)", SEC, SECFG)
        rows = [
            "Gerçekleşen brüt sonuç (TL)",
            "Yİ-ÜFE sonrası vergiye tabi sonuç (TL, negatif olabilir)",
            "Açık pozisyon - hâlâ elde tutulan adet",
            "Açık pozisyon - maliyet bedeli (TL)",
            "Durum",
        ]
    else:
        sect(ws, rb, 2, 14, "RESULTS FOR THIS ASSET (filled automatically by run)", SEC, SECFG)
        rows = [
            "Realised gross result (TL)",
            "Taxable result after YI-UFE (TL, can be negative)",
            "Open position - qty still held",
            "Open position - cost basis (TL)",
            "Status",
        ]
    status_lbl = rows[-1]  # last row is always the status row
    for i, lbl in enumerate(rows, rb + 1):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        lc = ws.cell(i, 2)
        lc.value = lbl
        lc.font = F(10, True)
        lc.alignment = AL()
        lc.border = B()
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=14)
        vc = ws.cell(i, 7)
        is_status = (lbl == status_lbl)
        vc.value = 0 if not is_status else ("Hesaplanmadı" if lang == "TR" else "Not run")
        vc.font = F(10, True)
        vc.alignment = AR() if not is_status else AL()
        vc.border = B()
        vc.fill = fill(POS if not is_status else WARN)
        if not is_status:
            vc.number_format = TLF if "qty" not in lbl.lower() and "adet" not in lbl.lower() else QTYF


def build_readme(wb, lang="TR"):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 105
    ws.row_dimensions[1].height = 30
    ws.merge_cells("B1:C1")
    t = ws["B1"]
    if lang == "TR":
        t.value = "NASIL KULLANILIR — LÜTFEN OKUYUN"
        lines = [
            ("GENEL BİLGİ", None),
            ("Bu dosyaya yurt dışı hisse senedi, ETF ve benzer menkul kıymet işlemlerinizi girin. Hesaplama web sitesinde yapılır.", CALC),
            ("Desteklenen para birimleri: USD (ABD Doları) ve EUR (Euro).", CALC),
            ("", None),
            ("NE DOLDURMANIZ GEREKİYOR", None),
            ("Mavi hücreleri doldurun: Tarih, İşlem Türü (ALIM veya SATIM), Adet, Fiyat (hisse başına), Komisyon.", CALC),
            ("Tarih formatı: GG.AA.YYYY veya YYYY-AA-GG (örn. 15.03.2025 veya 2025-03-15).", CALC),
            ("Fiyat ve komisyon, işlemin para birimiyle (USD veya EUR) girilmelidir.", CALC),
            ("Her varlık (hisse senedi veya ETF) için ayrı bir VARLIK sayfası kullanın.", CALC),
            ("", None),
            ("OTOMATİK HESAPLANAN ALANLAR", None),
            ("TCMB döviz kuru, önceki ay Yİ-ÜFE endeksi, TL tutarları, FIFO eşleme, endekslenmiş maliyet ve vergilendirilebilir sonuç otomatik hesaplanır.", CALC),
            ("Yİ-ÜFE verisi eksikse veya aylık endeks artışı %%10'un altındaysa çalışma kitabı TASLAK olarak işaretlenir.", CALC),
            ("", None),
            ("HESAPLAMAYI BAŞLATMAK", None),
            ("Bu dosyayı doldurduktan sonra web sitesine yükleyin ve 'Hesapla' butonuna basın.", CALC),
            ("Sonuçlar GİB Hazır Beyan formatında sunulur ve PDF olarak indirilebilir.", CALC),
            ("", None),
            ("TEMETTÜLER", None),
            ("Yurt dışı temettülerinizi web sitesindeki 'Temettü girişi' bölümünden ekleyin; bu dosyaya girmeyin.", CALC),
            ("Temettü geliri, menkul sermaye iradı kategorisindedir ve satış kazancından ayrı beyan edilir.", CALC),
            ("", None),
            ("SORUMLULUK REDDİ", None),
            ("Bu araç yalnızca tahmini hesaplama amaçlıdır; resmi vergi beyannamesi değildir. Beyan etmeden önce sonuçları bir mali müşavir ile teyit edin.", CALC),
        ]
    else:
        t.value = "HOW TO USE — PLEASE READ"
        lines = [
            ("OVERVIEW", None),
            ("Enter your foreign stock, ETF and similar securities trades in this file. The calculation is done on the website.", CALC),
            ("Supported currencies: USD (US Dollar) and EUR (Euro).", CALC),
            ("", None),
            ("WHAT YOU FILL IN", None),
            ("Fill the blue cells: Date, Type (BUY or SELL), Quantity, Price (per share), Commission.", CALC),
            ("Date formats accepted: DD.MM.YYYY or YYYY-MM-DD (e.g. 15.03.2025 or 2025-03-15).", CALC),
            ("Price and commission must be in the asset's currency (USD or EUR).", CALC),
            ("Use a separate ASSET sheet for each holding (stock or ETF).", CALC),
            ("", None),
            ("WHAT IS CALCULATED AUTOMATICALLY", None),
            ("TCMB exchange rate, prior-month Yİ-ÜFE inflation index, TL amounts, FIFO lot matching, indexed cost, and taxable result are all filled automatically.", CALC),
            ("If Yİ-ÜFE data is missing or the monthly index increase is below 10%, the workbook is marked PROVISIONAL.", CALC),
            ("", None),
            ("RUNNING THE CALCULATION", None),
            ("Upload this completed file on the website and press 'Calculate'.", CALC),
            ("Results are shown in GİB Hazır Beyan format and can be downloaded as a PDF.", CALC),
            ("", None),
            ("DIVIDENDS", None),
            ("Enter foreign dividends in the 'Dividend input' section on the website — not in this file.", CALC),
            ("Dividend income (menkul sermaye iradı) is declared separately from capital gains.", CALC),
            ("", None),
            ("DISCLAIMER", None),
            ("This tool provides estimates only and does not constitute an official tax filing. Verify all results with a licensed tax advisor before filing.", CALC),
        ]
    t.font = F(12, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = AC()
    for i, (txt, col) in enumerate(lines, 2):
        cell = ws.cell(i, 2)
        cell.value = txt
        header = col is None and txt != ""
        cell.font = F(10, header, WHITE if header else (col or CALC))
        if header:
            cell.fill = fill(NAVY)
        cell.alignment = AL()
    return ws


def build_dividend_guide_sheet(wb, lang="TR"):
    if "DIVIDEND_GUIDE" in wb.sheetnames:
        del wb["DIVIDEND_GUIDE"]
    ws = wb.create_sheet("DIVIDEND_GUIDE")
    ws.sheet_view.showGridLines = False
    for c, w in {"A": 3, "B": 34, "C": 95}.items():
        ws.column_dimensions[c].width = w
    ws.merge_cells("B1:C1")
    t = ws["B1"]
    if lang == "TR":
        t.value = "YURT DIŞI TEMETTÜ KILAVUZU - HENÜZ HESAP MAKINASI DEĞİL"
        rows = [
            ("Kapsam", "Yurt dışı hisse/ETF'yi şahsen tutan Türkiye mukimi bireyler. Ticari faaliyet, şirketler, türevler ve TEFAS fonları bu dosyanın dışındadır."),
            ("Kategori", "Yurt dışı hisse/ETF nakit dağıtımları genellikle menkul sermaye iradıdır; sermaye kazancı satış sonucu değildir."),
            ("Brüt tutara ulaşma", "Yabancı stopaj öncesi brüt temettüden başlayın. Yalnızca net tutar ve stopaj oranı biliniyorsa: brüt = net / (1 - stopaj oranı)."),
            ("TL'ye çevirme", "Brüt temettü ve yabancı vergiyi, ödeme/tahsil tarihindeki TCMB döviz alış kuruyla TL'ye çevirin. Kur kanıtını saklayın."),
            ("2026 beyan eşiği", "Stopaja tabi olmayan/istisna dışı menkul ve gayrimenkul sermaye iratları için 2026 eşiği 22.000 TL'dir. Aşılırsa tam tutar beyan edilir, sadece aşan kısım değil."),
            ("GVK 22/4 kontrolü", "Özel %50 istisna yalnızca GVK 22/4 koşulları sağlanırsa uygulanabilir. Küçük halka açık piyasa hissedarları genellikle önemli ortaklık koşulunu karşılamaz."),
            ("Yabancı vergi mahsubu", "Yabancı stopaj vergisi yalnızca yabancı kaynaklı gelire atfedilen Türk vergisi dahilinde ve belgeyle mahsup edilebilir."),
            ("Netleştirme yasağı", "Temettü geliri, hisse satış zararlarıyla netleştirilmez. Satış kazanç/zararları ile temettüler ayrı gelir kategorileridir."),
            ("İlerideki temettü çalışma kitabı", "Sütunlar şunları içermelidir: ödeme tarihi, ticker, ülke, döviz, brüt temettü, stopaj, net nakit, aracı komisyonu, TCMB kuru, brüt TL, yabancı vergi TL, GVK 22/4 bayrağı, vergiye tabi TL, notlar."),
            ("Resmi kaynaklar", "GVK güncel metni: https://www.mevzuat.gov.tr/mevzuat?MevzuatNo=193&MevzuatTertip=4&MevzuatTur=1"),
            ("Resmi kaynaklar", "GİB menkul sermaye iradı kılavuzu: https://intvrg.gib.gov.tr/hazirbeyan/assets/pdf/DUYURU_UNIVERSAL_2026_2026_menkulsermayeiradi.pdf"),
            ("Resmi kaynaklar", "GİB özet: https://intvrg.gib.gov.tr/hazirbeyan/menkulOzet.html"),
            ("Resmi kaynaklar", "TCMB kurları: https://www.tcmb.gov.tr/kurlar/kurlar_tr.html"),
        ]
        col_topic, col_rule = "Konu", "Kural / giriş notu"
    else:
        t.value = "FOREIGN DIVIDEND CALCULATION GUIDE - NOT A CALCULATOR YET"
        rows = [
            ("Scope", "Turkish tax resident individual holding foreign stocks/ETFs personally. Business trading, companies, derivatives, and TEFAS funds are outside this file."),
            ("Category", "Foreign stock/ETF cash distributions are normally menkul sermaye iradi, not a capital gain sale result."),
            ("Gross-up", "Start from gross dividend before foreign withholding. If only net cash and withholding rate are known: gross = net / (1 - withholding rate)."),
            ("TRY conversion", "Convert gross dividend and foreign tax to TL using TCMB forex buying rate for the payment/receipt date. Keep rate proof."),
            ("2026 declaration threshold", "For non-withheld/non-exempt menkul and real-estate capital income, the 2026 threshold is 22,000 TL. If exceeded, the full amount is declared, not only the excess."),
            ("GVK 22/4 check", "A special 50% exemption may apply only if GVK 22/4 conditions are met. Ordinary small public-market holdings normally do not meet the substantial ownership condition."),
            ("Foreign tax credit", "Foreign withholding tax can be credited only within Turkish tax attributable to the foreign-source income and only with documents."),
            ("No netting", "Dividend income is not netted against stock sale losses. Sale gains/losses and dividends are separate income categories."),
            ("Future dividend workbook", "Columns should include payment date, ticker, country, currency, gross dividend, withholding, net cash, broker fee, TCMB rate, gross TL, foreign tax TL, GVK 22/4 flag, taxable TL, and notes."),
            ("Official sources", "GVK current text: https://www.mevzuat.gov.tr/mevzuat?MevzuatNo=193&MevzuatTertip=4&MevzuatTur=1"),
            ("Official sources", "GIB menkul sermaye iradi guide: https://intvrg.gib.gov.tr/hazirbeyan/assets/pdf/DUYURU_UNIVERSAL_2026_2026_menkulsermayeiradi.pdf"),
            ("Official sources", "GIB summary: https://intvrg.gib.gov.tr/hazirbeyan/menkulOzet.html"),
            ("Official sources", "TCMB rates: https://www.tcmb.gov.tr/kurlar/kurlar_tr.html"),
        ]
        col_topic, col_rule = "Topic", "Rule / input note"
    t.font = F(12, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = AC()
    ws.cell(3, 2).value = col_topic
    ws.cell(3, 3).value = col_rule
    for cell in ws[3][1:3]:
        cell.font = F(10, True, WHITE)
        cell.fill = fill(NAVY)
        cell.alignment = AC()
        cell.border = B()
    for r, (topic, note) in enumerate(rows, 4):
        ws.cell(r, 2).value = topic
        ws.cell(r, 3).value = note
        for c in (2, 3):
            cell = ws.cell(r, c)
            cell.font = F(10, c == 2)
            cell.alignment = AL()
            cell.border = B()
            if c == 2:
                cell.fill = fill(SEC)
    ws.freeze_panes = "B4"
    return ws


def asset_sheet_names(wb):
    def key(name):
        m = re.search(r"(\d+)$", name)
        return (int(m.group(1)) if m else 10**9, name)
    return sorted([s for s in wb.sheetnames if s.startswith("ASSET_")], key=key)


def next_asset_index(wb):
    nums = []
    for name in asset_sheet_names(wb):
        m = re.search(r"(\d+)$", name)
        if m:
            nums.append(int(m.group(1)))
    return max(nums, default=0) + 1


def remove_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def build_summary(wb, totals=None, lang="TR"):
    totals = totals or {}
    remove_sheet_if_exists(wb, "SUMMARY")
    ws = wb.create_sheet("SUMMARY", 1 if "README" in wb.sheetnames else 0)
    ws.sheet_view.showGridLines = False
    for c, w in {"A": 3, "B": 22, "C": 10, "D": 18, "E": 22, "F": 18, "G": 18}.items():
        ws.column_dimensions[c].width = w
    ws.merge_cells("B1:G1")
    t = ws["B1"]
    if lang == "TR":
        t.value = f"YURT DIŞI MENKUL KIYMETLER - VERGİ ÖZETİ {TAX_YEAR}"
        heads = ["", "Varlık", "Döviz", "Brüt sonuç TL", "Vergiye tabi sonuç TL", "Açık adet", "Açık maliyet TL"]
    else:
        t.value = f"TURKEY FOREIGN SECURITIES - TAX SUMMARY {TAX_YEAR}"
        heads = ["", "Asset", "Currency", "Gross result TL", "Taxable result TL", "Open qty", "Open cost TL"]
    t.font = F(13, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = AC()

    _heads = heads  # alias used below
    for i, h in enumerate(_heads, 1):
        cell = ws.cell(3, i)
        cell.value = h
        cell.font = F(9, True, WHITE)
        cell.fill = fill(NAVY)
        cell.alignment = AC()
        cell.border = B()

    names = asset_sheet_names(wb) + option_sheet_names(wb)
    grand_gross = 0.0
    grand_raw = 0.0
    any_provisional = False
    any_incomplete = False
    all_notes = []

    for idx, name in enumerate(names, 4):
        d = totals.get(name, {})
        display = d.get("display_name")
        if display and display != name:
            asset_label = f"{display}  ({name})"
        elif display:
            asset_label = display
        else:
            # SUMMARY rebuilt without a run: read the name straight from the sheet
            if name.startswith("OPTION_"):
                sheet_name = get_option_name(wb[name]) if name in wb.sheetnames else name
            else:
                sheet_name = get_asset_name(wb[name]) if name in wb.sheetnames else name
            asset_label = f"{sheet_name}  ({name})" if sheet_name != name else name
        ws.cell(idx, 2).value = asset_label
        ws.cell(idx, 3).value = d.get("currency", "")
        ws.cell(idx, 4).value = d.get("gross", 0.0)
        ws.cell(idx, 5).value = d.get("taxable_raw", 0.0)
        ws.cell(idx, 6).value = round(d.get("open_qty", 0.0), 6)
        ws.cell(idx, 7).value = d.get("open_cost", 0.0)
        for c in range(2, 8):
            cell = ws.cell(idx, c)
            cell.font = F(10, False, LINK if c >= 4 else CALC)
            cell.alignment = AR() if c >= 4 else AL()
            cell.border = B()
            cell.fill = fill("F2F2F2")
            if c in (4, 5, 7):
                cell.number_format = TLF
            if c == 6:
                cell.number_format = QTYF
        grand_gross += d.get("gross", 0.0)
        grand_raw += d.get("taxable_raw", 0.0)
        any_provisional = any_provisional or d.get("provisional", False)
        any_incomplete = any_incomplete or d.get("incomplete", False)
        if d.get("note"):
            all_notes.append(f"{name}: {d['note']}")

    total_row = 4 + len(names)
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    tc = ws.cell(total_row, 2)
    tc.value = "YILLIK ZARAR TABAN ÖNCESI TOPLAM" if lang == "TR" else "TOTAL BEFORE ANNUAL LOSS FLOOR"
    tc.font = F(10, True, WHITE)
    tc.fill = fill(NAVY)
    tc.alignment = AR()
    tc.border = B()
    for c in range(4, 8):
        cell = ws.cell(total_row, c)
        if c == 4:
            cell.value = grand_gross
        elif c == 5:
            cell.value = grand_raw
        elif c == 6:
            cell.value = ""
        elif c == 7:
            cell.value = sum(totals.get(n, {}).get("open_cost", 0.0) for n in names)
        cell.font = F(10, True)
        cell.fill = fill(SEC)
        cell.alignment = AR()
        cell.border = B()
        if c in (4, 5, 7):
            cell.number_format = TLF

    tax_base = max(0.0, grand_raw)
    tax = income_tax(tax_base, TAX_YEAR)
    tb = total_row + 2
    if lang == "TR":
        sect(ws, tb, 2, 7, "GERÇEKLEŞMİŞ MENKUL KIYMET KAZANÇLARI ÜZERİNDEKİ GELİR VERGİSİ", NAVY)
        status = "KESİN"
        if any_incomplete:
            status = "EKSİK - bu dosyadan beyan yapmayın"
        elif any_provisional:
            status = "TASLAK - gerçekleşen bir satışta Yİ-ÜFE verisi eksik"
        items = [
            ("Zarar tabanı öncesi menkul kıymet vergiye tabi sonucu (TL)", grand_raw, TLF, None),
            ("Yıllık zarar tabanı sonrası nihai vergiye tabi matrah (TL)", tax_base, TLF, None),
            ("Bu matrah üzerindeki gelir vergisi (TL)", tax, TLF, None),
            ("1. taksit - Mart (TL)", tax / 2, TLF, POS),
            ("2. taksit - Temmuz (TL)", tax / 2, TLF, POS),
            ("Durum", status, None, WARN if status != "KESİN" else POS),
            ("Lot eşleme", "FIFO; aynı gün alımları gözden geçirilmeli", None, WARN),
            ("Temettüler", "Burada hesaplanmaz; DIVIDEND_GUIDE sayfasına bakın", None, WARN),
        ]
    else:
        sect(ws, tb, 2, 7, "INCOME TAX ON REALISED SECURITIES GAINS", NAVY)
        status = "FINAL"
        if any_incomplete:
            status = "INCOMPLETE - do not file from this workbook"
        elif any_provisional:
            status = "PROVISIONAL - missing YI-UFE data affected a realised sale"
        items = [
            ("Securities taxable result before loss floor (TL)", grand_raw, TLF, None),
            ("Final securities taxable base after annual loss floor (TL)", tax_base, TLF, None),
            ("Income tax due on this base (TL)", tax, TLF, None),
            ("1st instalment - March (TL)", tax / 2, TLF, POS),
            ("2nd instalment - July (TL)", tax / 2, TLF, POS),
            ("Status", status, None, WARN if status != "FINAL" else POS),
            ("Lot matching", "FIFO; same-day buys may require review", None, WARN),
            ("Dividends", "Not calculated here; see DIVIDEND_GUIDE", None, WARN),
        ]
    for i, (lbl, val, fmt, bg) in enumerate(items, tb + 1):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        lc = ws.cell(i, 2)
        lc.value = lbl
        lc.font = F(10, True)
        lc.fill = fill(NAVY if bg is None else bg)
        if bg is None:
            lc.font = F(10, True, WHITE)
        lc.alignment = AL()
        lc.border = B()
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=7)
        vc = ws.cell(i, 5)
        vc.value = val
        vc.font = F(10, True)
        vc.alignment = AR() if fmt else AL()
        vc.border = B()
        if fmt:
            vc.number_format = fmt
        if bg:
            vc.fill = fill(bg)

    if all_notes:
        nr = tb + len(items) + 2
        sect(ws, nr, 2, 7, "NOTLAR" if lang == "TR" else "NOTES", SEC, SECFG)
        for off, note in enumerate(all_notes[:25], 1):
            ws.merge_cells(start_row=nr + off, start_column=2, end_row=nr + off, end_column=7)
            cell = ws.cell(nr + off, 2)
            cell.value = note
            cell.font = F(9, True, "C00000")
            cell.alignment = AL()
    ws.freeze_panes = "B4"
    return ws


def build_transaction_view(wb, totals=None, lang="TR"):
    """Rebuild the TRANSACTION_VIEW sheet from scratch, newest first.
    `totals` is the dict returned by process_asset per sheet. If it is None
    or a sheet has no processed records, the view falls back to reading raw
    input rows so the sheet still reflects current data."""
    totals = totals or {}

    # Position the view right after SUMMARY (or after README) for consistency.
    if "SUMMARY" in wb.sheetnames:
        pos = wb.sheetnames.index("SUMMARY") + 1
    elif "README" in wb.sheetnames:
        pos = wb.sheetnames.index("README") + 1
    else:
        pos = 0
    remove_sheet_if_exists(wb, "TRANSACTION_VIEW")
    ws = wb.create_sheet("TRANSACTION_VIEW", pos)
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 3, "B": 11, "C": 14, "D": 12, "E": 5, "F": 9, "G": 11,
        "H": 11, "I": 12, "J": 9, "K": 11, "L": 20, "M": 15,
        "N": 14, "O": 14, "P": 34,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.row_dimensions[1].height = 26

    ws.merge_cells("B1:P1")
    t = ws["B1"]
    if lang == "TR":
        t.value = "İŞLEM GÖRÜNÜMÜ - EN YENİ ÖNCE (hesaplama sonrası otomatik yenilenir)"
        heads = [
            "Sayfa", "Varlık", "Tarih", "Sıra", "Tür", "Adet", "Fiyat",
            "Komisyon", "Döviz", "TCMB kuru", "Yİ-ÜFE ay/endeks",
            "TL tutar", "Brüt TL", "Vergiye tabi TL", "Durum",
        ]
    else:
        t.value = "TRANSACTION VIEW - NEWEST FIRST (rebuilt automatically by run)"
        heads = [
            "Sheet", "Asset", "Date", "Seq", "Type", "Quantity", "Price",
            "Commission", "Currency", "TCMB rate", "YI-UFE month/index",
            "TL amount", "Gross TL", "Taxable TL", "Status",
        ]
    t.font = F(12, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = AC()
    for ci, h in enumerate(heads, 2):
        c = ws.cell(3, ci)
        c.value = h
        c.font = F(9, True, WHITE)
        c.fill = fill(NAVY)
        c.alignment = AC()
        c.border = B()

    ws.merge_cells("B4:P4")
    note = ws.cell(4, 2)
    note.value = (
        "Bu görünüm her hesaplamada otomatik olarak yenilenir."
        if lang == "TR" else
        "This view is rebuilt automatically every time you run the calculation."
    )
    note.font = F(9, False, CALC)
    note.fill = fill(SEC)
    note.alignment = AL()

    # Collect all records across sheets.
    records = []
    for name in asset_sheet_names(wb):
        d = totals.get(name)
        if d and d.get("transactions"):
            records.extend(d["transactions"])
        else:
            # Fallback: read raw input rows so the view is never blank/stale.
            ws_a = wb[name]
            asset_name = get_asset_name(ws_a)
            currency = get_currency(ws_a)
            rb = find_result_row(ws_a)
            tx_last = max(TX_FIRST - 1, rb - 2)
            seq_by_date = {}
            for r in range(TX_FIRST, tx_last + 1):
                dd = parse_date(ws_a.cell(r, 3).value)
                typ = str(ws_a.cell(r, 4).value or "").strip().upper()
                qty = num(ws_a.cell(r, 5).value)
                if not dd or typ not in ("BUY", "SELL") or qty <= 0:
                    continue
                seq_by_date[dd] = seq_by_date.get(dd, 0) + 1
                records.append({
                    "sheet": name, "asset": asset_name, "date": dd,
                    "seq": seq_by_date[dd], "type": typ, "qty": qty,
                    "price": num(ws_a.cell(r, 6).value),
                    "comm": num(ws_a.cell(r, 7).value),
                    "currency": currency, "rate": None, "ufe_disp": "",
                    "tl_amount": None, "gross": None, "taxable": None,
                    "status": ("sadece giriş satırı - hesaplamak için hesaplamayı çalıştırın" if lang == "TR"
                              else "input row only - run script to calculate"),
                })

    for name in option_sheet_names(wb):
        d = totals.get(name)
        if d and d.get("transactions"):
            records.extend(d["transactions"])
        else:
            ws_o = wb[name]
            disp = get_option_name(ws_o)
            currency = get_option_currency(ws_o)
            rb = find_option_result_row(ws_o)
            tx_last = max(OPT_TX_FIRST - 1, rb - 2)
            seq_by_date = {}
            for r in range(OPT_TX_FIRST, tx_last + 1):
                dd = parse_date(ws_o.cell(r, 3).value)
                action = str(ws_o.cell(r, 4).value or "").strip().upper()
                if not dd or action not in OPT_ACTIONS:
                    continue
                seq_by_date[dd] = seq_by_date.get(dd, 0) + 1
                records.append({
                    "sheet": name, "asset": disp, "date": dd,
                    "seq": seq_by_date[dd], "type": action,
                    "qty": num(ws_o.cell(r, 5).value),
                    "price": num(ws_o.cell(r, 6).value),
                    "comm": num(ws_o.cell(r, 7).value),
                    "currency": currency, "rate": None, "ufe_disp": "",
                    "tl_amount": None, "gross": None, "taxable": None,
                    "status": ("sadece giriş satırı - hesaplamak için hesaplamayı çalıştırın" if lang == "TR"
                              else "input row only - run script to calculate"),
                })

    # Newest first: by date desc, then by seq desc within the same date.
    records.sort(key=lambda x: (x["date"], x["seq"]), reverse=True)

    row = 5
    for rec in records:
        typ = rec["type"]
        if typ in ("SELL", "STO", "STC"):
            bg = SELLBG
        elif typ in ("BUY", "BTO", "BTC"):
            bg = BUYBG
        else:                       # EXPIRE / EXERCISE / ASSIGN
            bg = "F2F2F2"
        values = [
            rec["sheet"], rec["asset"], rec["date"], rec["seq"], rec["type"],
            rec["qty"], rec["price"], rec["comm"], rec["currency"],
            rec.get("rate"), rec.get("ufe_disp", ""), rec.get("tl_amount"),
            rec.get("gross"), rec.get("taxable"), rec.get("status", ""),
        ]
        for ci, val in enumerate(values, 2):
            cell = ws.cell(row, ci)
            cell.value = val
            cell.font = F(9)
            cell.border = B()
            cell.fill = fill(bg)
            if ci == 4:               # Date
                cell.number_format = DATEF
                cell.alignment = AC()
            elif ci in (6,):          # Quantity
                cell.number_format = QTYF
                cell.alignment = AR()
            elif ci in (8, 9):        # Price, Commission
                cell.number_format = CUR
                cell.alignment = AR()
            elif ci == 11:            # TCMB rate
                cell.number_format = FXF
                cell.alignment = AR()
            elif ci in (13, 14, 15):  # TL amount, Gross, Taxable
                cell.number_format = TLF
                cell.alignment = AR()
            elif ci == 16:            # Status
                cell.alignment = AL()
            else:
                cell.alignment = AC()
        row += 1

    if not records:
        ws.merge_cells("B5:P5")
        c = ws.cell(5, 2)
        c.value = "No transactions yet. Enter BUY/SELL rows on the ASSET sheets, then run the script."
        c.font = F(9, False, CALC)
        c.alignment = AL()

    ws.freeze_panes = "B5"
    return ws


def build_blank_workbook(lang="TR", assets=None, rows=None, options=None):
    """Build a fresh blank workbook in memory and return it.
    Used by the web dashboard so it can pass the current UI language."""
    assets = assets or DEFAULT_ASSETS
    rows = rows or DEFAULT_MAX_TX
    options = options or DEFAULT_OPTIONS
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    build_readme(wb, lang=lang)
    for i in range(1, assets + 1):
        build_asset_sheet(wb, i, rows, lang=lang)
    for i in range(1, options + 1):
        build_option_sheet(wb, i, lang=lang)
    build_dividend_guide_sheet(wb, lang=lang)
    build_summary(wb, lang=lang)
    build_transaction_view(wb, lang=lang)
    return wb


def cmd_init(args):
    assets = int(get_option(args, "--assets", first_int(args, DEFAULT_ASSETS)))
    rows = int(get_option(args, "--rows", DEFAULT_MAX_TX))
    options = int(get_option(args, "--options", DEFAULT_OPTIONS))
    overwrite = "--overwrite" in args
    path = Path(FILENAME)
    if path.exists() and not overwrite:
        sys.exit(f"{FILENAME} already exists. Use add-asset, or rename it, or use --overwrite intentionally.")
    wb = build_blank_workbook(lang="TR", assets=assets, rows=rows, options=options)
    wb.save(FILENAME)
    print(f"Created {FILENAME} with {assets} asset sheets, {options} option sheets.")


def cmd_add_asset(args):
    count = first_int(args, default=1)
    rows = int(get_option(args, "--rows", DEFAULT_MAX_TX))
    try:
        wb = load_workbook(FILENAME)
    except FileNotFoundError:
        sys.exit(f"{FILENAME} not found. Run init first.")
    except PermissionError:
        sys.exit(f"Cannot open {FILENAME}. Close it in Excel first.")
    start = next_asset_index(wb)
    for idx in range(start, start + count):
        build_asset_sheet(wb, idx, rows)
    build_dividend_guide_sheet(wb)
    build_summary(wb)
    build_transaction_view(wb)
    try:
        wb.save(FILENAME)
    except PermissionError:
        sys.exit(f"Cannot save {FILENAME}. Close it in Excel first.")
    print(f"Added {count} asset sheet(s): ASSET_{start:02d} to ASSET_{start + count - 1:02d}.")


def cmd_add_option(args):
    count = first_int(args, default=1)
    try:
        wb = load_workbook(FILENAME)
    except FileNotFoundError:
        sys.exit(f"{FILENAME} not found. Run init first.")
    except PermissionError:
        sys.exit(f"Cannot open {FILENAME}. Close it in Excel first.")
    start = next_option_index(wb)
    for idx in range(start, start + count):
        build_option_sheet(wb, idx)
    build_summary(wb)
    build_transaction_view(wb)
    try:
        wb.save(FILENAME)
    except PermissionError:
        sys.exit(f"Cannot save {FILENAME}. Close it in Excel first.")
    print(f"Added {count} option sheet(s): OPTION_{start:02d} to OPTION_{start + count - 1:02d}.")


def cmd_upgrade(args):
    target_assets = int(get_option(args, "--assets", first_int(args, DEFAULT_ASSETS)))
    rows = int(get_option(args, "--rows", DEFAULT_MAX_TX))
    try:
        wb = load_workbook(FILENAME)
    except FileNotFoundError:
        sys.exit(f"{FILENAME} not found. Run init first.")
    except PermissionError:
        sys.exit(f"Cannot open {FILENAME}. Close it in Excel first.")

    if "README" not in wb.sheetnames:
        build_readme(wb)
    for name in asset_sheet_names(wb):
        ensure_asset_layout(wb[name])
    for name in option_sheet_names(wb):
        ensure_option_layout(wb[name])
    current = len(asset_sheet_names(wb))
    start = next_asset_index(wb)
    if current < target_assets:
        for idx in range(start, start + (target_assets - current)):
            build_asset_sheet(wb, idx, rows)
    if not option_sheet_names(wb):
        for i in range(1, DEFAULT_OPTIONS + 1):
            build_option_sheet(wb, i)
    build_dividend_guide_sheet(wb)
    build_summary(wb)
    build_transaction_view(wb)
    try:
        wb.save(FILENAME)
    except PermissionError:
        sys.exit(f"Cannot save {FILENAME}. Close it in Excel first.")
    print(f"Upgraded {FILENAME}. Asset sheets: {len(asset_sheet_names(wb))}, option sheets: {len(option_sheet_names(wb))}.")


# -----------------------------------------------------------------------------
# FETCHING AND PARSING
# -----------------------------------------------------------------------------
def tcmb_rate(date, currency, _cache={}, max_back=10):
    currency = (currency or "USD").strip().upper()
    if currency not in ("USD", "EUR"):
        currency = "USD"
    key = (date, currency)
    if key in _cache:
        return _cache[key]
    d = date
    for _ in range(max_back + 1):
        url = TCMB_XML.format(ym=d.strftime("%Y%m"), dmy=d.strftime("%d%m%Y"))
        try:
            r = requests.get(url, timeout=15)
            if r.status_code == 200 and r.content.strip().startswith(b"<?xml"):
                root = ET.fromstring(r.content)
                for cur in root.findall("Currency"):
                    if cur.get("CurrencyCode") == currency:
                        fb = cur.findtext("ForexBuying")
                        if fb and fb.strip():
                            res = (float(fb.replace(",", ".")), d)
                            _cache[key] = res
                            return res
        except (requests.RequestException, ET.ParseError, ValueError):
            pass
        d -= dt.timedelta(days=1)
    _cache[key] = (None, None)
    return (None, None)


def normalize_month_key(value):
    if value in (None, ""):
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m")
    s = str(value).strip()
    s = s.replace("/", "-").replace(".", "-")
    parts = [p for p in s.split("-") if p]
    try:
        if len(parts) == 2:
            if len(parts[0]) == 4:
                return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
            return f"{int(parts[1]):04d}-{int(parts[0]):02d}"
        if len(parts) >= 3:
            if len(parts[0]) == 4:
                return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
            return f"{int(parts[2]):04d}-{int(parts[1]):02d}"
    except ValueError:
        return None
    return None


def parse_evds_items(rows, series):
    out = {}
    series_keys = {series, series.replace(".", "_"), series.replace(".", "").upper(), series.upper()}
    for row in rows:
        if not isinstance(row, dict):
            continue
        ym = row.get("Tarih") or row.get("tarih") or row.get("DATE") or row.get("Date")
        month_key = normalize_month_key(ym)
        if not month_key:
            continue
        val = None
        for k in list(row.keys()):
            if k in series_keys or k.upper() in series_keys:
                val = row.get(k)
                break
        if val in (None, "", "null"):
            for k, v in row.items():
                if k not in ("Tarih", "tarih", "DATE", "Date", "UNIXTIME") and v not in (None, "", "null"):
                    val = v
                    break
        try:
            out[month_key] = float(str(val).replace(",", "."))
        except (ValueError, TypeError):
            continue
    return out


def evds_ufe(api_key, start="01-01-2018", end=None):
    if not api_key:
        return None
    if end is None:
        end = dt.date.today().strftime("%d-%m-%Y")
    headers = {"key": api_key, "User-Agent": "Mozilla/5.0"}
    for series in EVDS_YIUFE_CANDIDATES:
        for tmpl in EVDS_ENDPOINTS:
            url = tmpl.format(series=series, start=start, end=end)
            try:
                r = requests.get(url, headers=headers, timeout=20)
                r.raise_for_status()
                raw = r.text.strip()
                if not raw or raw[0] not in "[{":
                    continue
                data = r.json()
            except (requests.RequestException, ValueError):
                continue
            rows = data.get("items", data) if isinstance(data, dict) else data
            if not isinstance(rows, list) or not rows:
                continue
            out = parse_evds_items(rows, series)
            if out:
                print(f"  EVDS {series}: {len(out)} monthly YI-UFE values.")
                return out
    print("  EVDS failed for all YI-UFE series candidates; using manual UFE_TABLE.")
    return None


def evds_ufe_cached(api_key, start="01-01-2018", end=None):
    """Wrapper around evds_ufe that caches results once per day per API key.
    The cache is stored in .evds_cache.json next to this file.
    Each user's key is stored under a short hash — never the key itself."""
    if not api_key:
        return None
    today = dt.date.today().isoformat()
    key_hash = hashlib.sha256(api_key.encode()).hexdigest()[:16]

    cache = {}
    try:
        if EVDS_CACHE_FILE.exists():
            cache = json.loads(EVDS_CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        cache = {}

    entry = cache.get(key_hash, {})
    if entry.get("date") == today and entry.get("data"):
        print("  EVDS: returning cached Yİ-ÜFE data (already fetched today).")
        return entry["data"]

    result = evds_ufe(api_key, start, end)
    if result:
        cache[key_hash] = {"date": today, "data": result}
        try:
            EVDS_CACHE_FILE.write_text(
                json.dumps(cache, ensure_ascii=False), encoding="utf-8"
            )
        except Exception:
            pass
    return result


def prev_month_key(date):
    first = date.replace(day=1)
    prev = first - dt.timedelta(days=1)
    return prev.strftime("%Y-%m")


def ufe_value_for_key(month_key, ufe_map):
    if not ufe_map:
        return None
    return ufe_map.get(month_key)


# -----------------------------------------------------------------------------
# TAX ENGINE
# -----------------------------------------------------------------------------
def income_tax(net, year=TAX_YEAR):
    if net <= 0:
        return 0.0
    brackets = BRACKETS_BY_YEAR.get(year)
    if not brackets:
        raise ValueError(f"No tax brackets configured for {year}")
    prev = 0.0
    for upper, rate, base in brackets:
        if net <= upper:
            return base + (net - prev) * rate
        prev = upper
    return 0.0


def parse_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    for f in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


def num(v):
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def find_result_row(ws):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str):
            vu = v.upper()
            # Match both languages: TR sheets never contain the English
            # marker, so the old English-only check always fell through to
            # the fallback row number below on TR workbooks. That fallback
            # only happens to be correct when tx_rows == DEFAULT_MAX_TX; for
            # any other sheet size (custom --rows, add-asset, trades-built
            # workbooks) it points past the sheet into the results block's
            # own merged cells, and clearing "auto" columns there raises
            # AttributeError: 'MergedCell' object attribute 'value' is
            # read-only. Recognising the TR marker too fixes both the crash
            # and the (silent, guaranteed) English-labels-on-a-TR-sheet bug.
            if "RESULTS FOR THIS ASSET" in vu or "SONUÇLAR" in vu:
                return r
    return TX_FIRST + DEFAULT_MAX_TX + 1


def ensure_asset_layout(ws, lang="TR"):
    write_asset_headers(ws, lang=lang)
    rb = find_result_row(ws)
    if rb > ws.max_row or not ws.cell(rb, 2).value:
        write_asset_results_block(ws, rb, lang=lang)
    else:
        if lang == "TR":
            labels = [
                "Gerçekleşen brüt sonuç (TL)",
                "Yİ-ÜFE sonrası vergiye tabi sonuç (TL, negatif olabilir)",
                "Açık pozisyon - hâlâ elde tutulan adet",
                "Açık pozisyon - maliyet bedeli (TL)",
                "Durum",
            ]
        else:
            labels = [
                "Realised gross result (TL)",
                "Taxable result after YI-UFE (TL, can be negative)",
                "Open position - qty still held",
                "Open position - cost basis (TL)",
                "Status",
            ]
        # Only fill in labels that are genuinely missing (e.g. an older
        # workbook). Previously this unconditionally overwrote every label
        # with the English text, so every TR-language calculation silently
        # flipped its own results section to English on every run.
        for i, lbl in enumerate(labels, rb + 1):
            if not ws.cell(i, 2).value:
                ws.cell(i, 2).value = lbl
    return rb


def get_currency(ws):
    currency = "USD"
    for r in range(4, 9):
        label = ws.cell(r, 2).value
        if label and "currency" in str(label).lower():
            currency = str(ws.cell(r, 4).value or "USD").strip().upper()
            break
    if currency not in ("USD", "EUR"):
        currency = "USD"
    return currency


def get_asset_name(ws):
    """Read the user-entered asset name / ticker from the ASSET DETAILS block.
    Falls back to the sheet name if the cell is blank."""
    for r in range(4, 9):
        label = ws.cell(r, 2).value
        if label and "asset name" in str(label).lower():
            val = ws.cell(r, 4).value
            if val not in (None, ""):
                return str(val).strip()
            break
    return ws.title


def _consume_lots(remaining, pool, sell_ufe, sell_ufe_key):
    """Consume `remaining` shares from `pool` (a FIFO list, oldest at index 0).
    Returns (remaining_left, matched_cost, indexed_cost, missing_index, missing_months).
    Applies YI-UFE indexation when the index rose >= 10% from the lot's
    acquisition month to the sale month."""
    matched = 0.0
    indexed = 0.0
    missing = False
    miss_months = set()
    while remaining > 1e-9 and pool:
        lot = pool[0]
        take = min(remaining, lot["qty"])
        base = take * lot["unit_cost"]
        matched += base
        idx_base = base
        if lot.get("ufe") and sell_ufe and lot["ufe"] > 0:
            increase = (sell_ufe - lot["ufe"]) / lot["ufe"]
            if increase >= 0.10:
                idx_base = base * (sell_ufe / lot["ufe"])
        else:
            missing = True
            if not lot.get("ufe"):
                miss_months.add(lot.get("ufe_key"))
            if not sell_ufe:
                miss_months.add(sell_ufe_key)
        indexed += idx_base
        lot["qty"] -= take
        remaining -= take
        if lot["qty"] <= 1e-9:
            pool.pop(0)
    return remaining, matched, indexed, missing, miss_months


# =============================================================================
# OPTIONS MODULE  (foreign exchange-traded options -> deger artisi kazanci,
# GVK Muk. 80/81, same category as stocks (foreign exchange-traded).)
# =============================================================================
OPT_ACTIONS = ("BTO", "STO", "BTC", "STC", "EXPIRE", "EXERCISE", "ASSIGN")


def option_sheet_names(wb):
    def key(name):
        m = re.search(r"(\d+)$", name)
        return (int(m.group(1)) if m else 10**9, name)
    return sorted([s for s in wb.sheetnames if s.startswith("OPTION_")], key=key)


def next_option_index(wb):
    nums = []
    for name in option_sheet_names(wb):
        m = re.search(r"(\d+)$", name)
        if m:
            nums.append(int(m.group(1)))
    return max(nums, default=0) + 1


def write_option_headers(ws, lang="TR"):
    if lang == "TR":
        heads = [
            "#", "Tarih",
            "İşlem\n(BTO/STO/\nBTC/STC/\nEXPİRE/\nEXERCİSE/\nASSİGN)",
            "Kontrat",
            "Prim\n(hisse başına,\nvarlık dövizi)",
            "Komisyon\n(varlık dövizi,\nboş = 0)",
            "TCMB kuru\n(otomatik)",
            "Yİ-ÜFE\nönceki ay\n(otomatik)",
            "TL değer\n(otomatik)",
            "Gerçekleşen\nbrüt TL\n(otomatik)",
            "Gerçekleşen\nvergiye tabi TL\n(otomatik)",
            "Durum / not\n(otomatik)",
        ]
    else:
        heads = [
            "#", "Date",
            "Action\n(BTO/STO/\nBTC/STC/\nEXPIRE/\nEXERCISE/\nASSIGN)",
            "Contracts",
            "Premium\n(per share,\nasset ccy)",
            "Commission\n(asset ccy,\nblank = 0)",
            "TCMB rate\n(auto)",
            "YI-UFE index\nprev month\n(auto)",
            "TL value\n(auto)",
            "Realised\ngross TL\n(auto)",
            "Realised\ntaxable TL\n(auto)",
            "Status / note\n(auto)",
        ]
    for ci, h in enumerate(heads, 2):
        c = ws.cell(13, ci)
        c.value = h
        c.font = F(9, True, WHITE)
        c.fill = fill(NAVY)
        c.alignment = AC()
        c.border = B()


def write_option_results_block(ws, rb, lang="TR"):
    if lang == "TR":
        sect(ws, rb, 2, 13, "BU OPSİYON İÇİN SONUÇLAR (hesaplama sonrası otomatik doldurulur)", SEC, SECFG)
        rows = [
            "Gerçekleşen brüt sonuç (TL)",
            "Yİ-ÜFE sonrası vergiye tabi sonuç (TL, negatif olabilir)",
            "Açık pozisyon - hâlâ açık kontrat sayısı",
            "Açık pozisyon - net prim bedeli (TL)",
            "Durum",
        ]
    else:
        sect(ws, rb, 2, 13, "RESULTS FOR THIS OPTION (filled automatically by run)", SEC, SECFG)
        rows = [
            "Realised gross result (TL)",
            "Taxable result after YI-UFE (TL, can be negative)",
            "Open position - contracts still open",
            "Open position - net premium basis (TL)",
            "Status",
        ]
    status_lbl = rows[-1]
    for i, lbl in enumerate(rows, rb + 1):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        lc = ws.cell(i, 2)
        lc.value = lbl
        lc.font = F(10, True)
        lc.alignment = AL()
        lc.border = B()
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=13)
        vc = ws.cell(i, 7)
        is_status = (lbl == status_lbl)
        vc.value = 0 if not is_status else ("Hesaplanmadı" if lang == "TR" else "Not run")
        vc.font = F(10, True)
        vc.alignment = AR() if not is_status else AL()
        vc.border = B()
        vc.fill = fill(POS if not is_status else WARN)
        if not is_status:
            vc.number_format = TLF if "kontrat" not in lbl.lower() and "contracts" not in lbl.lower() else QTYF


def build_option_sheet(wb, idx, tx_rows=OPT_DEFAULT_ROWS, lang="TR"):
    ws = wb.create_sheet(f"OPTION_{idx:02d}")
    ws.sheet_view.showGridLines = False
    widths = {
        "A": 3, "B": 5, "C": 13, "D": 12, "E": 11, "F": 13, "G": 13,
        "H": 13, "I": 15, "J": 16, "K": 16, "L": 16, "M": 30,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B1:M1")
    t = ws["B1"]
    if lang == "TR":
        t.value = f"OPSİYON {idx:02d} - TEK SERİ, İŞLEMLER VE VERGİ (GVK Mük. 80/81)"
        details_header = "OPSİYON BİLGİLERİ (sarı hücreleri doldurun)"
        details = [
            ("Dayanak varlık ticker", ""),
            ("Hak (CALL veya PUT)", "CALL"),
            ("Kullanım fiyatı", ""),
            ("Vade tarihi", ""),
            ("Çarpan (kontrat başına hisse)", 100),
            ("Para birimi (USD veya EUR)", "USD"),
            ("Vergi yılı", TAX_YEAR),
        ]
        tx_header = "İŞLEMLER — her ayağı tarih sırasıyla girin (BTO/STO aç, STC/BTC kapat, EXPIRE süresi doldu)"
    else:
        t.value = f"OPTION {idx:02d} - ONE SERIES, TRANSACTIONS AND TAX (GVK Muk. 80/81)"
        details_header = "OPTION DETAILS (fill the yellow cells)"
        details = [
            ("Underlying ticker", ""),
            ("Right (CALL or PUT)", "CALL"),
            ("Strike price", ""),
            ("Expiry date", ""),
            ("Multiplier (shares per contract)", 100),
            ("Currency (USD or EUR)", "USD"),
            ("Tax year", TAX_YEAR),
        ]
        tx_header = "TRANSACTIONS - enter each leg in date order (BTO/STO open, STC/BTC close, EXPIRE if it lapsed)"
    t.font = F(12, True, WHITE)
    t.fill = fill(NAVY)
    t.alignment = AC()

    sect(ws, 3, 2, 13, details_header, SEC, SECFG)
    for i, (lbl, val) in enumerate(details, OPT_DETAIL_FIRST):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        lc = ws.cell(i, 2)
        lc.value = lbl
        lc.font = F(10, True)
        lc.alignment = AL()
        lc.border = B()
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=7)
        vc = ws.cell(i, 5)
        vc.value = val
        vc.font = F(10, False, INP)
        vc.fill = fill(WARN)
        vc.alignment = AL()
        vc.border = B()

    sect(ws, 12, 2, 13, tx_header, NAVY)
    write_option_headers(ws, lang=lang)

    for r in range(OPT_TX_FIRST, OPT_TX_FIRST + tx_rows):
        ws.cell(r, 2).value = r - OPT_TX_FIRST + 1
        ws.cell(r, 2).font = F(9)
        ws.cell(r, 2).alignment = AC()
        ws.cell(r, 2).border = B()
        for c in range(3, 14):
            cell = ws.cell(r, c)
            cell.border = B()
            cell.alignment = AR()
            if c in (3, 4, 5, 6, 7):
                cell.fill = fill("EBF3FB")
                cell.font = F(9, False, INP)
                if c == 3:
                    cell.number_format = DATEF
                elif c == 4:
                    cell.alignment = AC()
                elif c == 5:
                    cell.number_format = QTYF
                else:
                    cell.number_format = CUR
            elif c in (8, 9):
                cell.fill = fill("FFF7E6")
                cell.font = F(9, False, CALC)
                cell.number_format = FXF if c == 8 else '#,##0.00'
            elif c == 13:
                cell.fill = fill("F2F2F2")
                cell.font = F(9, False, CALC)
                cell.alignment = AL()
            else:
                cell.fill = fill("F2F2F2")
                cell.font = F(9, False, LINK)
                cell.number_format = TLF

    rb = OPT_TX_FIRST + tx_rows + 1
    write_option_results_block(ws, rb, lang=lang)
    ws.freeze_panes = "B14"
    return ws


def find_option_result_row(ws):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str):
            vu = v.upper()
            if "RESULTS FOR THIS OPTION" in vu or "SONUÇLAR" in vu:
                return r
    return OPT_TX_FIRST + OPT_DEFAULT_ROWS + 1


def ensure_option_layout(ws, lang="TR"):
    write_option_headers(ws, lang=lang)
    rb = find_option_result_row(ws)
    if rb > ws.max_row or not ws.cell(rb, 2).value:
        write_option_results_block(ws, rb, lang=lang)
    return rb


def _opt_detail(ws, needle):
    for r in range(OPT_DETAIL_FIRST, OPT_DETAIL_FIRST + 7):
        label = ws.cell(r, 2).value
        if label and needle in str(label).lower():
            return ws.cell(r, 5).value
    return None


def get_option_currency(ws):
    cur = str(_opt_detail(ws, "currency") or "USD").strip().upper()
    return cur if cur in ("USD", "EUR") else "USD"


def get_option_multiplier(ws):
    m = num(_opt_detail(ws, "multiplier"))
    return m if m and m > 0 else 100.0


def get_option_name(ws):
    under = _opt_detail(ws, "underlying")
    right = _opt_detail(ws, "right")
    strike = _opt_detail(ws, "strike")
    expiry = _opt_detail(ws, "expiry")
    parts = []
    if under not in (None, ""):
        parts.append(str(under).strip())
    if right not in (None, ""):
        parts.append(str(right).strip().upper())
    if strike not in (None, ""):
        parts.append(str(strike).strip())
    exp = parse_date(expiry)
    if exp:
        parts.append(exp.strftime("%Y-%m-%d"))
    return " ".join(parts) if parts else ws.title


def process_option_sheet(ws, ufe_map, lang="TR"):
    rb = ensure_option_layout(ws, lang=lang)
    tx_last = max(OPT_TX_FIRST - 1, rb - 2)
    currency = get_option_currency(ws)
    mult = get_option_multiplier(ws)
    disp = get_option_name(ws)
    underlying = str(_opt_detail(ws, "underlying") or "").strip()
    print(f"  Option: {disp}  ({currency}, x{mult:g})")

    # Clear auto columns.
    for r in range(OPT_TX_FIRST, tx_last + 1):
        for c in range(8, 14):
            ws.cell(r, c).value = None

    txs = []
    order = 0
    for r in range(OPT_TX_FIRST, tx_last + 1):
        d = parse_date(ws.cell(r, 3).value)
        action = str(ws.cell(r, 4).value or "").strip().upper()
        contracts = num(ws.cell(r, 5).value)
        premium = num(ws.cell(r, 6).value)
        comm = num(ws.cell(r, 7).value)
        if not d or action not in OPT_ACTIONS:
            continue
        if action not in ("EXPIRE",) and contracts <= 0:
            continue
        txs.append({"row": r, "date": d, "action": action,
                    "contracts": contracts, "premium": premium,
                    "comm": comm, "order": order})
        order += 1
    txs.sort(key=lambda t: (t["date"], t["order"]))

    lots = []            # open lots, each {side, shares, unit_val, ufe, ufe_key, date}
    gross_total = 0.0
    taxable_raw = 0.0
    incomplete = False
    provisional = False
    exercise_present = False
    has_close = False
    missing_index_months = set()
    notes = []
    tx_records = []

    def emit(t, tl_value, gross, taxable, status):
        ufe = t.get("ufe")
        ufe_key = t.get("ufe_key")
        if ufe is not None:
            ufe_disp = f"{ufe_key} / {ufe:g}"
        elif ufe_key:
            ufe_disp = f"{ufe_key} / (missing)"
        else:
            ufe_disp = ""
        tx_records.append({
            "sheet": ws.title, "asset": disp, "date": t["date"],
            "seq": t["order"] + 1, "type": t["action"], "qty": t["contracts"],
            "price": t["premium"], "comm": t["comm"], "currency": currency,
            "rate": t.get("rate"), "ufe_disp": ufe_disp,
            "tl_amount": tl_value, "gross": gross, "taxable": taxable,
            "status": status,
        })

    for t in txs:
        rate, used_date = tcmb_rate(t["date"], currency)
        t["rate"] = rate
        ws.cell(t["row"], 8).value = round(rate, 4) if rate else None
        ufe_key = prev_month_key(t["date"])
        t["ufe_key"] = ufe_key
        u = ufe_value_for_key(ufe_key, ufe_map)
        t["ufe"] = u
        ws.cell(t["row"], 9).value = u if u is not None else None

        action = t["action"]
        shares = t["contracts"] * mult
        if not rate and action != "EXPIRE":
            incomplete = True
            missing_rate_status = ("TCMB kuru eksik - tarihi kontrol edin" if lang == "TR"
                                    else "missing TCMB rate - check date")
            notes.append((f"{t['date']} {currency} için TCMB kuru bulunamadı" if lang == "TR"
                          else f"No TCMB rate for {t['date']} {currency}"))
            ws.cell(t["row"], 13).value = missing_rate_status
            emit(t, None, None, None, missing_rate_status)
            continue

        # ---- OPENING legs ----
        if action == "BTO":   # buy to open: long, premium paid = indexable cost
            cost = (t["premium"] * shares + t["comm"]) * rate
            lots.append({"side": "LONG", "shares": shares,
                         "unit_val": cost / shares if shares else 0.0,
                         "ufe": u, "ufe_key": ufe_key, "date": t["date"]})
            ws.cell(t["row"], 10).value = cost
            long_opened_status = "uzun pozisyon açıldı" if lang == "TR" else "long opened"
            ws.cell(t["row"], 13).value = long_opened_status
            emit(t, cost, None, None, long_opened_status)
            continue
        if action == "STO":   # sell to open: short, premium received = proceeds
            credit = (t["premium"] * shares - t["comm"]) * rate
            lots.append({"side": "SHORT", "shares": shares,
                         "unit_val": credit / shares if shares else 0.0,
                         "ufe": None, "ufe_key": ufe_key, "date": t["date"]})
            ws.cell(t["row"], 10).value = credit
            short_opened_status = ("kısa pozisyon açıldı (prim alındı)" if lang == "TR"
                                    else "short opened (premium received)")
            ws.cell(t["row"], 13).value = short_opened_status
            emit(t, credit, None, None, short_opened_status)
            continue

        # ---- EXERCISE / ASSIGN: premium rolls into the stock, no option P&L ----
        if action in ("EXERCISE", "ASSIGN"):
            exercise_present = True
            need = shares if shares > 0 else sum(l["shares"] for l in lots)
            closed = 0.0
            while need > 1e-9 and lots:
                lot = lots[0]
                take = min(need, lot["shares"])
                lot["shares"] -= take
                need -= take
                closed += take
                if lot["shares"] <= 1e-9:
                    lots.pop(0)
            tl_info = (t["premium"] * (closed) * (rate or 0.0)) if rate else None
            und = underlying or ("dayanak varlık" if lang == "TR" else "underlying")
            if lang == "TR":
                msg = f"{action}: primi {und} ASSET sayfasına aktarın (opsiyon K/Z olarak sayılmaz)"
                notes.append(f"{t['date']} tarihli {action}: {und} hisse sayfasını elle güncelleyin")
            else:
                msg = f"{action}: roll premium into ASSET sheet for {und} (not counted as option P&L)"
                notes.append(f"{action} on {t['date']}: adjust the {und} stock sheet manually")
            ws.cell(t["row"], 10).value = tl_info
            ws.cell(t["row"], 13).value = msg
            emit(t, tl_info, None, None, msg)
            continue

        # ---- CLOSING legs: STC closes LONG, BTC closes SHORT, EXPIRE closes all ----
        has_close = True
        if action == "EXPIRE":
            need = shares if shares > 0 else sum(l["shares"] for l in lots)
            leg_unit_close = 0.0     # nothing paid/received at expiry
            close_is_proceeds = None  # depends on lot side
        elif action == "STC":
            need = shares
            close_tl = (t["premium"] * shares - t["comm"]) * rate
            leg_unit_close = close_tl / shares if shares else 0.0
            close_is_proceeds = True
            ws.cell(t["row"], 10).value = close_tl
        elif action == "BTC":
            need = shares
            close_tl = (t["premium"] * shares + t["comm"]) * rate
            leg_unit_close = close_tl / shares if shares else 0.0
            close_is_proceeds = False
            ws.cell(t["row"], 10).value = close_tl
        else:
            continue

        leg_gross = 0.0
        leg_taxable = 0.0
        leg_missing = False
        matched_any = False
        while need > 1e-9 and lots:
            lot = lots[0]
            # Validate the close direction matches the lot side (except EXPIRE).
            if action == "STC" and lot["side"] != "LONG":
                break
            if action == "BTC" and lot["side"] != "SHORT":
                break
            take = min(need, lot["shares"])
            matched_any = True
            if lot["side"] == "LONG":
                cost = take * lot["unit_val"]
                proceeds = take * (leg_unit_close if action != "EXPIRE" else 0.0)
                idx_cost = cost
                if lot.get("ufe") and u and lot["ufe"] > 0:
                    inc = (u - lot["ufe"]) / lot["ufe"]
                    if inc >= 0.10:
                        idx_cost = cost * (u / lot["ufe"])
                else:
                    leg_missing = True
                    if not lot.get("ufe"):
                        missing_index_months.add(lot.get("ufe_key"))
                    if not u:
                        missing_index_months.add(ufe_key)
                leg_gross += proceeds - cost
                leg_taxable += proceeds - idx_cost
            else:  # SHORT lot
                credit = take * lot["unit_val"]            # premium received at open
                cost = take * (leg_unit_close if action != "EXPIRE" else 0.0)
                leg_gross += credit - cost
                leg_taxable += credit - cost                # shorts: no indexation
            lot["shares"] -= take
            need -= take
            if lot["shares"] <= 1e-9:
                lots.pop(0)

        if need > 1e-9 and action != "EXPIRE":
            incomplete = True
            notes.append(f"{action} on {t['date']} exceeds open contracts by {need / mult:g}")

        gross_total += leg_gross
        taxable_raw += leg_taxable
        if leg_missing:
            provisional = True
        ws.cell(t["row"], 11).value = leg_gross
        ws.cell(t["row"], 12).value = leg_taxable
        if action == "EXPIRE":
            if lang == "TR":
                st = "değersiz sona erdi" if matched_any else "EXPIRE ama açık pozisyon yok"
            else:
                st = "expired worthless" if matched_any else "EXPIRE but nothing open"
        elif leg_missing:
            st = (f"{action} kapatıldı (taslak - Yİ-ÜFE eksik)" if lang == "TR"
                  else f"{action} closed (provisional - missing YI-UFE)")
        else:
            st = f"{action} kapatıldı" if lang == "TR" else f"{action} closed"
        ws.cell(t["row"], 13).value = st
        emit(t, ws.cell(t["row"], 10).value, leg_gross, leg_taxable, st)

    open_contracts = sum(l["shares"] for l in lots) / mult if mult else 0.0
    net_basis = sum((l["unit_val"] * l["shares"]) * (1 if l["side"] == "LONG" else -1) for l in lots)

    ws.cell(rb + 1, 7).value = gross_total
    ws.cell(rb + 2, 7).value = taxable_raw
    ws.cell(rb + 3, 7).value = round(open_contracts, 6)
    ws.cell(rb + 4, 7).value = net_basis

    if incomplete:
        status_code = "INCOMPLETE"
        status = ("EKSİK - eksik döviz kuru veya fazla kapatılmış kontratları düzeltin" if lang == "TR"
                  else "INCOMPLETE - fix missing FX or over-closed contracts")
    elif provisional:
        status_code = "PROVISIONAL"
        months = ", ".join(sorted(m for m in missing_index_months if m))
        status = (f"TASLAK - Yİ-ÜFE eksik: {months}" if lang == "TR"
                  else f"PROVISIONAL - missing YI-UFE for: {months}")
    elif exercise_present:
        status_code = "ACTION_NEEDED"
        status = ("İŞLEM GEREKLİ - kullanım/devir: hisse sayfasını güncelleyin" if lang == "TR"
                  else "ACTION NEEDED - exercise/assignment: adjust stock sheet")
    elif has_close:
        status_code = "FINAL"
        status = "Mevcut verilerle KESİN" if lang == "TR" else "FINAL from available data"
    else:
        status_code = "NONE"
        status = "Gerçekleşen opsiyon sonucu yok" if lang == "TR" else "No realised option results"
    ws.cell(rb + 5, 7).value = status
    ws.cell(rb + 5, 7).alignment = AL()
    ws.cell(rb + 5, 7).fill = fill(WARN if (incomplete or provisional or exercise_present) else POS)

    if notes:
        note = " | ".join(dict.fromkeys(notes))
        ws.cell(rb + 7, 2).value = note[:1000]
        ws.cell(rb + 7, 2).font = F(9, True, "C00000")
        ws.cell(rb + 7, 2).alignment = AL()

    return {
        "currency": currency,
        "display_name": disp,
        "transactions": tx_records,
        "gross": gross_total,
        "taxable_raw": taxable_raw,
        "open_qty": open_contracts,
        "open_cost": net_basis,
        "provisional": provisional,
        "incomplete": incomplete,
        "note": " | ".join(dict.fromkeys(notes)) if notes else "",
        "status_code": status_code,
        "is_option": True,
    }


def process_asset(ws, ufe_map, lang="TR"):
    rb = ensure_asset_layout(ws, lang=lang)
    tx_last = max(TX_FIRST - 1, rb - 2)
    currency = get_currency(ws)
    asset_name = get_asset_name(ws)
    print(f"  Currency: {currency}")

    # Clear old calculation columns.
    for r in range(TX_FIRST, tx_last + 1):
        for c in range(8, 15):
            ws.cell(r, c).value = None

    txs = []
    order = 0
    for r in range(TX_FIRST, tx_last + 1):
        d = parse_date(ws.cell(r, 3).value)
        typ = str(ws.cell(r, 4).value or "").strip().upper()
        # The TR column header and README both instruct users to type
        # ALIM/SATIM, but the engine only ever matched literal BUY/SELL -
        # every row a TR user typed as instructed was silently skipped
        # (no error, just a quietly wrong zero-gains result). Normalize
        # both spellings so TR input actually gets processed.
        if typ in ("ALIM", "AL"):
            typ = "BUY"
        elif typ in ("SATIM", "SAT"):
            typ = "SELL"
        qty = num(ws.cell(r, 5).value)
        price = num(ws.cell(r, 6).value)
        comm = num(ws.cell(r, 7).value)
        if not d or typ not in ("BUY", "SELL") or qty <= 0:
            continue
        txs.append({
            "row": r, "date": d, "type": typ, "qty": qty,
            "price": price, "comm": comm, "order": order,
        })
        order += 1
    # Sort by date, then by execution/sheet order within the day.
    # We no longer force all BUYs before all SELLs: same-day ordering is
    # handled explicitly by the same-day matching rule in the FIFO phase.
    txs.sort(key=lambda t: (t["date"], t["order"]))

    incomplete = False
    notes = []
    open_lot_missing_ufe = set()

    for t in txs:
        rate, used_date = tcmb_rate(t["date"], currency)
        t["rate"] = rate
        t["rate_date"] = used_date
        ws.cell(t["row"], 8).value = round(rate, 4) if rate else None
        idx_key = prev_month_key(t["date"])
        t["ufe_key"] = idx_key
        u = ufe_value_for_key(idx_key, ufe_map)
        t["ufe"] = u
        ws.cell(t["row"], 9).value = u if u is not None else None
        if not rate:
            incomplete = True
            notes.append((f"{t['date']} {currency} için TCMB kuru bulunamadı" if lang == "TR"
                          else f"No TCMB rate for {t['date']} {currency}"))
            print(f"    {t['date']} {t['type']:<4} ! no TCMB rate found")
            continue
        if t["type"] == "BUY":
            t["tl_cost"] = (t["qty"] * t["price"] + t["comm"]) * rate
            ws.cell(t["row"], 10).value = t["tl_cost"]
            if u is None:
                open_lot_missing_ufe.add(idx_key)
        else:
            t["tl_proceeds"] = (t["qty"] * t["price"] - t["comm"]) * rate
            ws.cell(t["row"], 10).value = t["tl_proceeds"]
        flag = "" if used_date == t["date"] else f" (rate from {used_date})"
        print(f"    {t['date']} {t['type']:<4} {t['qty']:g} @ {t['price']:g} {currency}/TRY {rate:.4f}{flag}")

    main_lots = []           # older FIFO pool, oldest at index 0
    gross_total = 0.0
    taxable_raw = 0.0
    provisional = False
    missing_index_months = set()
    has_sales = False
    tx_records = []          # rows for the TRANSACTION_VIEW sheet

    def ufe_disp_for(t):
        uk, u = t.get("ufe_key"), t.get("ufe")
        if u is not None:
            return f"{uk} / {u:g}"
        return f"{uk} / (missing)" if uk else ""

    # Process day by day so same-day ordering is handled explicitly. Within a
    # day, ALL of the day's buys are made available to the day's sells
    # regardless of which row comes first, so the result is order-independent.
    for day, day_iter in groupby(txs, key=lambda t: t["date"]):
        day_txs = list(day_iter)
        seq = 0

        # Phase 1: gather this day's BUYs into a same-day pool.
        sameday_lots = []
        for t in day_txs:
            if t["type"] != "BUY":
                continue
            seq += 1
            t["seq"] = seq
            missing_rate_status = ("TCMB kuru eksik - tarihi kontrol edin" if lang == "TR"
                                    else "missing TCMB rate - check date")
            if not t.get("rate"):
                tx_records.append({
                    "sheet": ws.title, "asset": asset_name, "date": t["date"],
                    "seq": seq, "type": "BUY", "qty": t["qty"],
                    "price": t["price"], "comm": t["comm"], "currency": currency,
                    "rate": None, "ufe_disp": ufe_disp_for(t), "tl_amount": None,
                    "gross": None, "taxable": None,
                    "status": missing_rate_status,
                })
                continue
            unit_cost = t["tl_cost"] / t["qty"] if t["qty"] else 0.0
            sameday_lots.append({
                "qty": t["qty"], "unit_cost": unit_cost,
                "ufe": t["ufe"], "ufe_key": t["ufe_key"], "date": t["date"],
            })
            tx_records.append({
                "sheet": ws.title, "asset": asset_name, "date": t["date"],
                "seq": seq, "type": "BUY", "qty": t["qty"],
                "price": t["price"], "comm": t["comm"], "currency": currency,
                "rate": t.get("rate"), "ufe_disp": ufe_disp_for(t),
                "tl_amount": t["tl_cost"],
                "gross": None, "taxable": None,
                "status": "alım lotu kaydedildi" if lang == "TR" else "buy lot recorded",
            })

        # STRICT FIFO: today's buys join the main pool now (as the newest lots,
        # appended at the end), so a same-day sell still consumes the OLDEST
        # shares first. The same-day pool is then empty, so nothing is matched
        # ahead of older lots. (Order-independence is preserved because all of
        # today's buys are added before any of today's sells are processed.)
        if not SAME_DAY_MATCHING:
            main_lots.extend(sameday_lots)
            sameday_lots = []

        # Phase 2: process this day's SELLs in execution order.
        for t in day_txs:
            if t["type"] != "SELL":
                continue
            seq += 1
            t["seq"] = seq
            if not t.get("rate"):
                tx_records.append({
                    "sheet": ws.title, "asset": asset_name, "date": t["date"],
                    "seq": seq, "type": "SELL", "qty": t["qty"],
                    "price": t["price"], "comm": t["comm"], "currency": currency,
                    "rate": None, "ufe_disp": ufe_disp_for(t), "tl_amount": None,
                    "gross": None, "taxable": None,
                    "status": ("TCMB kuru eksik - tarihi kontrol edin" if lang == "TR"
                              else "missing TCMB rate - check date"),
                })
                continue

            has_sales = True
            remaining = t["qty"]
            matched_cost = 0.0
            indexed_cost = 0.0
            sell_missing_index = False

            # Same-day pool first (empty under strict FIFO), then the main pool.
            before_sd = remaining
            remaining, m, ix, miss, mm = _consume_lots(
                remaining, sameday_lots, t.get("ufe"), t.get("ufe_key"))
            used_sameday = before_sd - remaining
            matched_cost += m
            indexed_cost += ix
            sell_missing_index = sell_missing_index or miss
            missing_index_months |= mm

            remaining, m, ix, miss, mm = _consume_lots(
                remaining, main_lots, t.get("ufe"), t.get("ufe_key"))
            matched_cost += m
            indexed_cost += ix
            sell_missing_index = sell_missing_index or miss
            missing_index_months |= mm

            oversold = remaining > 1e-9
            if oversold:
                incomplete = True
                notes.append((f"{t['date']} tarihli SATIM, mevcut ALIM lotlarını {remaining:g} adet aşıyor" if lang == "TR"
                              else f"SELL on {t['date']} exceeds available BUY lots by {remaining:g} shares"))

            proceeds = t.get("tl_proceeds", 0.0)
            gross = proceeds - matched_cost
            taxable = proceeds - indexed_cost
            gross_total += gross
            taxable_raw += taxable
            if sell_missing_index:
                provisional = True
            ws.cell(t["row"], 11).value = matched_cost
            ws.cell(t["row"], 12).value = indexed_cost
            ws.cell(t["row"], 13).value = gross
            ws.cell(t["row"], 14).value = taxable

            if lang == "TR":
                if oversold:
                    sell_status = "SATIM mevcut lotları aşıyor"
                elif sell_missing_index:
                    sell_status = "SATIM hesaplandı (taslak - Yİ-ÜFE eksik)"
                elif used_sameday > 1e-9:
                    sell_status = f"SATIM hesaplandı (aynı gün eşleşen {used_sameday:g})"
                else:
                    sell_status = "SATIM hesaplandı"
            else:
                if oversold:
                    sell_status = "SELL exceeds available lots"
                elif sell_missing_index:
                    sell_status = "SELL calculated (provisional - missing YI-UFE)"
                elif used_sameday > 1e-9:
                    sell_status = f"SELL calculated (same-day matched {used_sameday:g})"
                else:
                    sell_status = "SELL calculated"
            tx_records.append({
                "sheet": ws.title, "asset": asset_name, "date": t["date"],
                "seq": seq, "type": "SELL", "qty": t["qty"],
                "price": t["price"], "comm": t["comm"], "currency": currency,
                "rate": t.get("rate"), "ufe_disp": ufe_disp_for(t),
                "tl_amount": proceeds,
                "gross": gross, "taxable": taxable,
                "status": sell_status,
            })

        # Phase 3 (same-day rule only): any same-day buys not consumed today
        # join the main pool, preserving execution order, for future days.
        # Under strict FIFO this list is already empty.
        for lot in sameday_lots:
            if lot["qty"] > 1e-9:
                main_lots.append(lot)

    lots = main_lots

    open_qty = sum(l["qty"] for l in lots)
    open_cost = sum(l["qty"] * l["unit_cost"] for l in lots)

    ws.cell(rb + 1, 7).value = gross_total
    ws.cell(rb + 2, 7).value = taxable_raw
    ws.cell(rb + 3, 7).value = round(open_qty, 6)
    ws.cell(rb + 4, 7).value = open_cost

    if incomplete:
        status_code = "INCOMPLETE"
        status = ("EKSİK - eksik döviz kuru veya eşleşmemiş satışları düzeltin" if lang == "TR"
                  else "INCOMPLETE - fix missing FX or unmatched sells")
    elif provisional:
        status_code = "PROVISIONAL"
        months = ", ".join(sorted(m for m in missing_index_months if m))
        status = (f"TASLAK - gerçekleşen satış ayları için Yİ-ÜFE eksik: {months}" if lang == "TR"
                  else f"PROVISIONAL - missing YI-UFE for realised sale months: {months}")
        notes.append(status)
    elif has_sales:
        status_code = "FINAL"
        status = "Mevcut verilerle KESİN" if lang == "TR" else "FINAL from available data"
    else:
        status_code = "NONE"
        status = "Gerçekleşen satış yok" if lang == "TR" else "No realised sales"
    if open_lot_missing_ufe and not provisional:
        notes.append(
            ("Gelecekteki satışlar için açık lotlarda Yİ-ÜFE eksik olan aylar: " if lang == "TR"
             else "Open lots have missing YI-UFE months for future sales: ")
            + ", ".join(sorted(open_lot_missing_ufe)))
    ws.cell(rb + 5, 7).value = status
    ws.cell(rb + 5, 7).alignment = AL()
    ws.cell(rb + 5, 7).fill = fill(WARN if (incomplete or provisional) else POS)

    if notes:
        note = " | ".join(dict.fromkeys(notes))
        ws.cell(rb + 7, 2).value = note[:1000]
        ws.cell(rb + 7, 2).font = F(9, True, "C00000")
        ws.cell(rb + 7, 2).alignment = AL()
    else:
        ws.cell(rb + 7, 2).value = None

    return {
        "currency": currency,
        "display_name": asset_name,
        "transactions": tx_records,
        "gross": gross_total,
        "taxable_raw": taxable_raw,
        "open_qty": open_qty,
        "open_cost": open_cost,
        "provisional": provisional,
        "incomplete": incomplete,
        "note": " | ".join(dict.fromkeys(notes)) if notes else "",
        "status_code": status_code,
    }


def build_workbook_from_trades(trades):
    """Build an in-memory Turkey_Tax_Tracker workbook from a flat trade list.
    `trades`: list of dicts {date, asset, action, qty, price, currency, commission}.
    One ASSET sheet per (asset, currency). Returns the Workbook (not saved)."""
    from collections import OrderedDict
    groups = OrderedDict()
    for t in trades:
        key = (str(t.get("asset", "UNKNOWN")).upper(),
               str(t.get("currency", "USD")).upper())
        groups.setdefault(key, []).append(t)

    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)

    for idx, ((asset, currency), rows) in enumerate(groups.items(), 1):
        n = max(DEFAULT_MAX_TX, len(rows) + 5)
        ws = build_asset_sheet(wb, idx, n)
        ws.cell(4, 4).value = asset
        ws.cell(5, 4).value = currency if currency in ("USD", "EUR") else "USD"
        for r, t in enumerate(sorted(rows, key=lambda x: x["date"]), TX_FIRST):
            ws.cell(r, 3).value = t["date"]
            ws.cell(r, 4).value = str(t.get("action", "")).upper()
            ws.cell(r, 5).value = t.get("qty", 0)
            ws.cell(r, 6).value = t.get("price", 0)
            ws.cell(r, 7).value = t.get("commission", 0) or 0

    build_dividend_guide_sheet(wb)
    build_summary(wb)
    build_transaction_view(wb)
    return wb


def run_calculation(wb, evds_key=None, lang="TR"):
    """Process every ASSET and OPTION sheet in `wb`, rebuild SUMMARY and
    TRANSACTION_VIEW, and return a results dict. Does NOT save the file.

    Used by both the command line (cmd_run) and the web dashboard, so the two
    produce identical numbers from the same engine. Yi-UFE comes from EVDS when
    a key is supplied, otherwise the built-in UFE_TABLE.
    """
    if "README" not in wb.sheetnames:
        build_readme(wb, lang=lang)
    if "DIVIDEND_GUIDE" not in wb.sheetnames:
        build_dividend_guide_sheet(wb)

    ufe_map = evds_ufe_cached(evds_key) or UFE_TABLE
    if ufe_map is not UFE_TABLE:
        ufe_source = "EVDS"
    else:
        ufe_source = "yerleşik Yİ-ÜFE tablosu" if lang == "TR" else "built-in UFE_TABLE"

    totals = {}
    for name in asset_sheet_names(wb):
        totals[name] = process_asset(wb[name], ufe_map, lang=lang)
    for name in option_sheet_names(wb):
        totals[name] = process_option_sheet(wb[name], ufe_map, lang=lang)

    build_summary(wb, totals, lang=lang)
    build_transaction_view(wb, totals, lang=lang)

    grand_raw = sum(d["taxable_raw"] for d in totals.values())
    tax_base = max(0.0, grand_raw)
    tax = income_tax(tax_base, TAX_YEAR)

    incomplete = any(d.get("incomplete") for d in totals.values())
    provisional = any(d.get("provisional") for d in totals.values())
    if incomplete:
        status_code = "INCOMPLETE"
        status = ("EKSİK - işaretli satırlar düzeltilene kadar beyan etmeyin." if lang == "TR"
                  else "INCOMPLETE - do not file until the flagged rows are fixed.")
    elif provisional:
        status_code = "PROVISIONAL"
        status = ("TASLAK - eksik Yİ-ÜFE, gerçekleşen en az bir satışı etkiledi." if lang == "TR"
                  else "PROVISIONAL - missing YI-UFE affected at least one realised sale.")
    else:
        status_code = "FINAL"
        status = "Mevcut verilerle KESİN." if lang == "TR" else "FINAL from available data."

    # Flat per-instrument lines for display.
    lines = []
    for name in asset_sheet_names(wb) + option_sheet_names(wb):
        d = totals.get(name, {})
        if d.get("is_option"):
            kind = "Opsiyon" if lang == "TR" else "Option"
        else:
            kind = "Hisse/ETF" if lang == "TR" else "Stock/ETF"
        flag_code = ("INCOMPLETE" if d.get("incomplete")
                     else "PROVISIONAL" if d.get("provisional") else "OK")
        lines.append({
            "sheet": name,
            "name": d.get("display_name", name),
            "kind": kind,
            "currency": d.get("currency", ""),
            "gross": d.get("gross", 0.0),
            "taxable_raw": d.get("taxable_raw", 0.0),
            "open_qty": d.get("open_qty", 0.0),
            "open_cost": d.get("open_cost", 0.0),
            "note": d.get("note", ""),
            "flag": flag_code,
            "status_code": d.get("status_code", flag_code),
        })

    warnings = [f"{ln['name']}: {ln['note']}" for ln in lines if ln["note"]]

    return {
        "totals": totals,
        "lines": lines,
        "raw_result": grand_raw,
        "tax_base": tax_base,
        "estimated_tax": tax,
        "instalment_1": tax / 2.0,
        "instalment_2": tax / 2.0,
        "status": status,
        "status_code": status_code,
        "incomplete": incomplete,
        "provisional": provisional,
        "ufe_source": ufe_source,
        "warnings": warnings,
        "tax_year": TAX_YEAR,
    }


def cmd_run(args):
    evds_key = None
    for a in args:
        if not a.startswith("--"):
            evds_key = a
            break
    print("Loading workbook...")
    try:
        wb = load_workbook(FILENAME)
    except FileNotFoundError:
        sys.exit(f"{FILENAME} not found. Run: python tax_tool.py init")
    except PermissionError:
        sys.exit(f"Cannot open {FILENAME}. Close it in Excel first.")

    print("Fetching YI-UFE and processing all sheets...")
    res = run_calculation(wb, evds_key)
    print(f"  Yi-UFE source: {res['ufe_source']}")

    try:
        wb.save(FILENAME)
    except PermissionError:
        sys.exit(f"Cannot save {FILENAME}. Close it in Excel and re-run.")
    print(f"\nDone. Raw securities result: {res['raw_result']:,.2f} TL | "
          f"Tax base after annual loss floor: {res['tax_base']:,.2f} TL | "
          f"Tax due: {res['estimated_tax']:,.2f} TL")
    print(f"Status: {res['status']}")


# -----------------------------------------------------------------------------
# DIVIDEND GUIDE TEXT OUTPUT
# -----------------------------------------------------------------------------
# DIVIDEND GUIDE TEXT OUTPUT
# -----------------------------------------------------------------------------
def dividend_guide_text():
    lines = [
        "Turkey foreign dividend calculation guide",
        "",
        "Scope",
        "  This guide covers a Turkish tax resident individual holding foreign",
        "  stocks or ETFs personally (not via a company).",
        "",
        "Core rule",
        "  Foreign cash dividends are normally menkul sermaye iradi (MSI).",
        "  Do not net dividends against foreign stock sale losses.",
        "",
        "Step 1 - collect gross dividend",
        "  Use gross dividend before any foreign withholding tax.",
        "  gross = net / (1 - withholding_rate) if only net amount is known.",
        "",
        "Step 2 - convert to TL",
        "  Convert gross dividend and foreign tax withheld to TL at TCMB",
        "  forex buying rate (doviz alis) on the dividend payment date.",
        "",
        "Step 3 - check GVK 22/4 exemption",
        "  50 pct exemption only if GVK 22/4 conditions are met.",
        "  Ordinary small holdings typically do not qualify.",
        "",
        "Step 4 - apply the 2026 declaration threshold",
        "  Threshold for 2025 income year is 22,000 TL.",
        "  If the threshold is exceeded, declare the full MSI, not just the excess.",
        "",
        "Step 5 - compute Turkish income tax",
        "  Apply the progressive tariff for the relevant tax year to your total",
        "  taxable income including the MSI.",
        "",
        "Step 6 - foreign tax credit",
        "  Foreign tax withheld may be credited against Turkish tax on that",
        "  foreign income only, and only with supporting documents.",
        "",
        "Useful links",
        "  - GiB Hazir Beyan: https://hazirbeyan.gib.gov.tr",
        "  - TCMB rates: https://www.tcmb.gov.tr/kurlar/kurlar_tr.html",
    ]
    return "\n".join(lines)


def cmd_guide():
    out = Path("Dividend_Calculation_Guide.md")
    out.write_text(dividend_guide_text(), encoding="utf-8")
    print(f"Wrote {out}")


# -----------------------------------------------------------------------------
# CLI UTILS
# -----------------------------------------------------------------------------
def get_option(args, name, default=None):
    if name not in args:
        return default
    i = args.index(name)
    if i + 1 >= len(args):
        return default
    return args[i + 1]


def first_int(args, default=1):
    for a in args:
        if a.startswith("--"):
            continue
        try:
            return int(a)
        except ValueError:
            continue
    return default


def usage():
    print("Turkey Foreign Stock / ETF Tax Tool")
    print("Usage:")
    print("  python tax_tool.py init [N] [--assets N] [--rows N] [--options N] [--overwrite]")
    print("  python tax_tool.py upgrade [N] [--assets N] [--rows N]")
    print("  python tax_tool.py add-asset [N] [--rows N]")
    print("  python tax_tool.py add-option [N]")
    print("  python tax_tool.py run [EVDS_KEY]")
    print("  python tax_tool.py guide")


def main():
    if len(sys.argv) < 2:
        usage()
        return
    cmd = sys.argv[1]
    args = sys.argv[2:]
    if cmd == "init":
        cmd_init(args)
    elif cmd == "upgrade":
        cmd_upgrade(args)
    elif cmd in ("add-asset", "add_asset"):
        cmd_add_asset(args)
    elif cmd in ("add-option", "add_option"):
        cmd_add_option(args)
    elif cmd == "run":
        cmd_run(args)
    elif cmd == "guide":
        cmd_guide()
    else:
        usage()


if __name__ == "__main__":
    main()
