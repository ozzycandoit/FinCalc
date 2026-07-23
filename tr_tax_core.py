"""
tr_tax_core.py - calculation layer for the Streamlit dashboard.

Two entry points:

1) calculate_from_workbook(uploaded_file, evds_key)  -> THE FULL ENGINE.
   Runs the exact same code as the desktop tool (tax_tool.py): stocks AND
   options, Yi-UFE indexation, strict FIFO + same-day rule, USD/EUR,
   commissions, full-category netting, the official 2026 tariff and the
   March/July instalments. The web app and the workbook therefore produce
   identical numbers. Input is your Turkey_Tax_Tracker.xlsx.

2) calculate_turkish_taxes(uploaded_file) -> QUICK CHECK (stocks only).
   A lightweight per-asset FIFO on a flat broker statement, with NO Yi-UFE
   indexation and NO options. Handy for a fast sanity check; not a filing.
"""

import io
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET

import requests
from openpyxl import load_workbook

import tax_tool   # the full, verified engine


# =============================================================================
# GIB "Hazir Beyan" mapping  (turns engine results into declaration sections)
# =============================================================================
def build_gib_beyan(results, dividends=None):
    """Map engine results into GiB annual-return sections.

    3.D Diger Kazanc ve Irat  -> capital gains, GVK Muk. 80/1
    3.C Menkul Sermaye Iradi  -> foreign dividends (code 622), if provided

    Column meaning (mirrors the GiB Hazir Beyan summary):
      Gayrisafi = gross gain/income before deductions
      Gider/Indirim = expenses + Yi-UFE indexation relief
      Safi = net taxable figure that is declared
    """
    lines = results.get("lines", [])
    gross_nominal = sum(ln.get("gross", 0.0) for ln in lines)
    tax_base = results.get("tax_base", max(0.0, results.get("raw_result", 0.0)))

    gayrisafi_cg = max(0.0, gross_nominal)
    safi_cg = tax_base
    gider_cg = max(0.0, gayrisafi_cg - safi_cg)

    capital = {
        "code": "GVK Mük. 80/1",
        "label": "Yurt dışı menkul kıymet / türev değer artışı kazancı",
        "gayrisafi": gayrisafi_cg,
        "gider_indirim": gider_cg,
        "safi": safi_cg,
        "kesilen": 0.0,
    }

    div_section = None
    if dividends:
        g = sum(d.get("gross_tl", 0.0) for d in dividends)
        ind = sum(d.get("expense_tl", 0.0) for d in dividends)
        kes = sum(d.get("withheld_tl", 0.0) for d in dividends)
        div_section = {
            "code": "622 - Yurt dışından elde edilen diğer menkul sermaye iratları",
            "gayrisafi": g,
            "indirilecek": ind,
            "safi": max(0.0, g - ind),
            "kesilen": kes,
            "rows": dividends,
        }

    return {
        "capital_gains": capital,
        "dividends": div_section,
        "tax_base": tax_base,
        "tax": results.get("estimated_tax", 0.0),
        "instalment_1": results.get("instalment_1", 0.0),
        "instalment_2": results.get("instalment_2", 0.0),
        "tax_year": results.get("tax_year", 2026),
        "status": results.get("status", ""),
    }


def calculate_from_midas(pdf_files, dividend_entries=None, evds_key=None):
    """Full pipeline: parse Midas PDFs -> build workbook -> run full engine.
    Returns run_calculation results plus the parsed-trade preview and dividends
    parsed from the statements (merged with manual dividend_entries)."""
    import midas_pdf
    parsed = midas_pdf.parse_many(pdf_files)

    if parsed["trades"]:
        wb = tax_tool.build_workbook_from_trades(parsed["trades"])
        results = tax_tool.run_calculation(wb, evds_key or None)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        results["workbook_bytes"] = bio.getvalue()
    else:
        results = {
            "totals": {}, "lines": [], "raw_result": 0.0, "tax_base": 0.0,
            "estimated_tax": 0.0, "instalment_1": 0.0, "instalment_2": 0.0,
            "status": "İşlem bulunamadı (ekstrelerde alım/satım yok).",
            "ufe_source": "-", "warnings": [], "tax_year": 2026,
            "workbook_bytes": b"",
        }

    div_inputs = [{"date": d["date"], "currency": d["currency"],
                   "gross": d["gross"], "withholding": d["withholding"],
                   "expense": d.get("expense", 0.0)} for d in parsed["dividends"]]
    if dividend_entries:
        div_inputs.extend(dividend_entries)
    results["dividends_computed"] = compute_dividends(div_inputs) if div_inputs else None

    results["parsed_trades"] = parsed["trades"]
    results["parsed_dividends"] = parsed["dividends"]
    results["parse_warnings"] = parsed["warnings"]
    results["periods"] = parsed["periods"]
    return results


def calculate_from_trade_rows(trades, dividend_entries=None, evds_key=None):
    """Run the full engine on a confirmed/edited flat trade list (e.g. after the
    user reviews PDF-parsed rows). `trades`: {date, asset, action, qty, price,
    currency, commission}."""
    clean = []
    for t in trades:
        d = parse_flexible_date(t.get("date"))
        act = str(t.get("action", "")).strip().upper()
        qty = _num(t.get("qty"))
        if d is None or act not in ("BUY", "SELL") or qty <= 0:
            continue
        ccy = str(t.get("currency", "USD")).strip().upper()
        clean.append({"date": d, "asset": str(t.get("asset", "UNKNOWN")).upper(),
                      "action": act, "qty": qty, "price": _num(t.get("price")),
                      "currency": ccy if ccy in ("USD", "EUR") else "USD",
                      "commission": _num(t.get("commission"))})
    if not clean:
        return {"totals": {}, "lines": [], "raw_result": 0.0, "tax_base": 0.0,
                "estimated_tax": 0.0, "instalment_1": 0.0, "instalment_2": 0.0,
                "status": "Geçerli işlem yok.", "ufe_source": "-", "warnings": [],
                "tax_year": 2026, "workbook_bytes": b""}
    wb = tax_tool.build_workbook_from_trades(clean)
    results = tax_tool.run_calculation(wb, evds_key or None)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    results["workbook_bytes"] = bio.getvalue()
    results["dividends_computed"] = compute_dividends(dividend_entries) if dividend_entries else None
    return results


def detailed_transactions(results):
    """Flat, date-sorted list of every processed leg for the detail report."""
    rows = []
    for name, d in results.get("totals", {}).items():
        for t in d.get("transactions", []):
            rows.append(t)
    rows.sort(key=lambda x: (x.get("date") or "", x.get("seq", 0)))
    return rows


def compute_dividends(entries):
    """entries: list of dicts {date, currency, gross, withholding, expense}.
    Returns a list with TL conversions added (TCMB doviz alis on the date)."""
    out = []
    for e in entries:
        d = parse_flexible_date(e.get("date"))
        if d is None:
            continue
        ccy = (str(e.get("currency") or "USD").upper())
        if ccy not in ("USD", "EUR"):
            ccy = "USD"
        gross = _num(e.get("gross"))
        wh = _num(e.get("withholding"))
        exp = _num(e.get("expense"))
        if gross <= 0:
            continue
        rate, _u = get_tcmb_rate(d, ccy)
        if rate is None:
            continue
        out.append({
            "date": d, "currency": ccy, "rate": rate,
            "gross_tl": gross * rate,
            "withheld_tl": wh * rate,
            "expense_tl": exp * rate,
        })
    return out


# =============================================================================
# 1) FULL ENGINE  (everything, identical to the desktop tool)
# =============================================================================
def calculate_from_workbook(uploaded_file, evds_key=None, lang="TR"):
    """Process an uploaded Turkey_Tax_Tracker.xlsx through the real engine.

    Returns the dict from tax_tool.run_calculation plus 'workbook_bytes', the
    processed workbook (rates, gains, SUMMARY, TRANSACTION_VIEW filled in) ready
    to download.
    """
    wb = load_workbook(uploaded_file)
    results = tax_tool.run_calculation(wb, evds_key or None, lang=lang)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    results["workbook_bytes"] = bio.getvalue()
    return results


# =============================================================================
# 2) QUICK CHECK  (flat broker statement, stocks only, no indexation)
# =============================================================================
BRACKETS_2026 = [
    (190000, 0.15, 0),
    (400000, 0.20, 28500),
    (1000000, 0.27, 70500),
    (5300000, 0.35, 232500),
    (float("inf"), 0.40, 1737500),
]


def get_tcmb_rate(trade_date, currency="USD", max_back=10):
    """Official TCMB 'doviz alis' rate; walks back for weekends/holidays.
    Returns (rate, used_date) or (None, None) - never a fabricated fallback."""
    currency = (currency or "USD").upper()
    current = trade_date
    for _ in range(max_back):
        if current.date() == datetime.today().date():
            url = "https://www.tcmb.gov.tr/kurlar/today.xml"
        else:
            url = (f"https://www.tcmb.gov.tr/kurlar/{current.strftime('%Y%m')}/"
                   f"{current.strftime('%d%m%Y')}.xml")
        try:
            resp = requests.get(url, timeout=5)
            if resp.status_code == 200:
                tree = ET.fromstring(resp.content)
                for cur in tree.findall("Currency"):
                    if cur.get("CurrencyCode") == currency:
                        node = cur.find("ForexBuying")
                        if node is not None and node.text:
                            return float(node.text), current
        except Exception:
            pass
        current -= timedelta(days=1)
    return None, None


def parse_flexible_date(date_val):
    if isinstance(date_val, datetime):
        return date_val
    if not isinstance(date_val, str):
        return None
    s = date_val.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def income_tax_2026(base):
    if base <= 0:
        return 0.0
    prev = 0.0
    for upper, rate, cumulative in BRACKETS_2026:
        if base <= upper:
            return cumulative + (base - prev) * rate
        prev = upper
    return 0.0


def _num(v):
    if v in (None, ""):
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def build_quick_template_bytes(lang="TR"):
    """Blank one-sheet flat workbook for the free Quick Check tier.
    Much simpler than the full Turkey_Tax_Tracker.xlsx: one row per trade,
    no options, no Yi-UFE. Returns xlsx bytes ready for st.download_button.

    Styled to match the app's editorial look (Claude palette): a title band,
    a real instruction note, a bordered header row, a grayed-out example row,
    dropdown validation on the Action/Currency columns so users can't typo
    them, and light borders on every data cell so it reads as a table even
    with gridlines hidden."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "QUICK_CHECK"
    ws.sheet_view.showGridLines = False

    ACCENT = "D97757"   # Claude terracotta
    INK    = "3D3929"   # Claude dark text
    CREAM  = "FAF9F5"   # Claude background
    LINE   = "E5E1D6"   # hairline border
    MUTED  = "8A8578"   # muted gray for hints/example row

    if lang == "TR":
        title = "HIZLI KONTROL — Hisse / ETF İşlemleri"
        subtitle = "Opsiyon yok · Yİ-ÜFE yok · sadece hızlı bir tahmin"
        headers = ["Tarih", "Varlık", "İşlem", "Adet", "Fiyat",
                   "Döviz", "Komisyon (ops.)"]
        note = ("Her satıra bir ALIM veya SATIM işlemi girin. Tarih: GG.AA.YYYY veya "
                "YYYY-AA-GG. 2. satır bir örnektir — silip kendi verinizle değiştirin. "
                "İşlem ve Döviz hücrelerinde açılır listeden seçim yapabilirsiniz.")
        action_choices = ["ALIM", "SATIM"]
        example = ["15.01.2025", "AAPL", "ALIM", 10, 150, "USD", 1]
    else:
        title = "QUICK CHECK — Stock / ETF Trades"
        subtitle = "No options · no Yİ-ÜFE indexation · a fast estimate only"
        headers = ["Date", "Asset", "Action", "Quantity", "Price",
                   "Currency", "Commission (opt.)"]
        note = ("Enter one BUY or SELL per row. Date: DD.MM.YYYY or YYYY-MM-DD. "
                "Row 2 is an example — delete it and replace with your own data. "
                "Use the dropdowns in the Action and Currency cells.")
        action_choices = ["BUY", "SELL"]
        example = ["15.01.2025", "AAPL", "BUY", 10, 150, "USD", 1]

    thin = Side(style="thin", color=LINE)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title band ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = title
    t.font = Font(name="Georgia", size=14, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=ACCENT)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = subtitle
    s.font = Font(name="Arial", size=10, italic=True, color="FFFFFF")
    s.fill = PatternFill("solid", fgColor=ACCENT)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # ── Instruction note ─────────────────────────────────────────────────────
    ws.merge_cells("A3:G4")
    n = ws["A3"]
    n.value = note
    n.font = Font(name="Arial", size=9, color=MUTED)
    n.fill = PatternFill("solid", fgColor=CREAM)
    n.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    widths = [14, 14, 12, 10, 12, 10, 16]
    for col, w in zip("ABCDEFG", widths):
        ws.column_dimensions[col].width = w

    # ── Header row ───────────────────────────────────────────────────────────
    header_row = 5
    for ci, h in enumerate(headers, 1):
        c = ws.cell(header_row, ci)
        c.value = h
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = cell_border
    ws.row_dimensions[header_row].height = 24

    # ── Example row (grayed out, italic — a visible sample, not real data) ──
    example_row = header_row + 1
    for ci, v in enumerate(example, 1):
        c = ws.cell(example_row, ci)
        c.value = v
        c.font = Font(name="Arial", size=10, italic=True, color=MUTED)
        c.fill = PatternFill("solid", fgColor="F0EEE5")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border

    # ── Blank data rows with a visible grid (borders, not just a flat fill) ─
    first_data_row = example_row + 1
    last_data_row = first_data_row + 60
    for r in range(first_data_row, last_data_row):
        for c in range(1, 8):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor=CREAM)
            cell.border = cell_border
        ws.cell(r, 4).number_format = "0.####"     # Adet / Quantity
        ws.cell(r, 5).number_format = "#,##0.00"   # Fiyat / Price
        ws.cell(r, 7).number_format = "#,##0.00"   # Komisyon / Commission

    # ── Dropdown validation: Action column, Currency column ─────────────────
    dv_action = DataValidation(type="list", formula1=f'"{",".join(action_choices)}"',
                                allow_blank=True, showDropDown=False)
    dv_action.error = "Please choose a value from the dropdown."
    dv_action.errorTitle = "Invalid entry"
    ws.add_data_validation(dv_action)
    dv_action.add(f"C{example_row}:C{last_data_row - 1}")

    dv_ccy = DataValidation(type="list", formula1='"USD,EUR"',
                             allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_ccy)
    dv_ccy.add(f"F{example_row}:F{last_data_row - 1}")

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:G{header_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def calculate_turkish_taxes(uploaded_file):
    """Quick per-asset FIFO on a flat statement. Columns:
    Date | Asset | Action(BUY/SELL) | Quantity | Price | [Currency] | [Commission]
    Header row optional (auto-detected). No Yi-UFE indexation, no options."""
    wb = load_workbook(uploaded_file, data_only=True)
    sheet = wb.active

    trades, rate_cache, warnings = [], {}, []
    all_rows = list(sheet.iter_rows(values_only=True))

    start = 0
    if all_rows:
        first = all_rows[0]
        d0 = parse_flexible_date(first[0]) if len(first) > 0 else None
        a0 = (str(first[2]).strip().upper() if len(first) > 2 and first[2] else "")
        if d0 is None and a0 not in ("BUY", "SELL", "ALIM", "AL", "SATIM", "SAT"):
            start = 1

    for row in all_rows[start:]:
        if not row or row[0] is None:
            continue
        trade_date = parse_flexible_date(row[0])
        if trade_date is None:
            warnings.append(f"Skipped a row with an unreadable date: {row[0]!r}")
            continue
        asset = (str(row[1]).strip().upper() if len(row) > 1 and row[1] else "UNKNOWN")
        action = (str(row[2]).strip().upper() if len(row) > 2 and row[2] else "")
        if action in ("ALIM", "AL"):
            action = "BUY"
        elif action in ("SATIM", "SAT"):
            action = "SELL"
        if action not in ("BUY", "SELL"):
            continue
        qty = _num(row[3]) if len(row) > 3 else 0.0
        price = _num(row[4]) if len(row) > 4 else 0.0
        currency = (str(row[5]).strip().upper() if len(row) > 5 and row[5] else "USD")
        if currency not in ("USD", "EUR"):
            currency = "USD"
        commission = _num(row[6]) if len(row) > 6 else 0.0
        if qty <= 0:
            continue
        ck = (trade_date.date(), currency)
        if ck not in rate_cache:
            rate_cache[ck] = get_tcmb_rate(trade_date, currency)
        rate, _u = rate_cache[ck]
        if rate is None:
            warnings.append(f"No TCMB {currency} rate near {trade_date.date()} - {asset} row excluded.")
            continue
        trades.append({"date": trade_date, "asset": asset, "action": action,
                       "qty": qty, "price": price, "commission": commission,
                       "currency": currency, "rate": rate})

    trades.sort(key=lambda t: t["date"])
    queues, total_gain_tl = {}, 0.0
    for t in trades:
        q = queues.setdefault((t["asset"], t["currency"]), [])
        if t["action"] == "BUY":
            cost = (t["qty"] * t["price"] + t["commission"]) * t["rate"]
            q.append({"qty": t["qty"], "unit_cost_tl": cost / t["qty"]})
        else:
            proceeds = (t["qty"] * t["price"] - t["commission"]) * t["rate"]
            unit_p = proceeds / t["qty"] if t["qty"] else 0.0
            rem = t["qty"]
            while rem > 1e-9 and q:
                lot = q[0]
                take = min(rem, lot["qty"])
                total_gain_tl += take * (unit_p - lot["unit_cost_tl"])
                lot["qty"] -= take
                rem -= take
                if lot["qty"] <= 1e-9:
                    q.pop(0)
            if rem > 1e-9:
                warnings.append(f"{t['asset']}: SELL on {t['date'].date()} exceeds available BUY lots by {rem:g}.")

    base = max(0.0, total_gain_tl)
    tax = income_tax_2026(base)
    return {
        "total_gains": base, "raw_result": total_gain_tl,
        "estimated_tax": tax, "instalment_1": tax / 2.0, "instalment_2": tax / 2.0,
        "trades_processed": len(trades), "indexation_applied": False,
        "warnings": warnings,
        "status": "Success" if not warnings else "Completed with warnings",
    }
